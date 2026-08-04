"""Treasury Account Symbol component structure and FAST Book edition capture.

The Bureau of the Fiscal Service documents the Treasury Account Symbol (TAS)
as eight named component fields ("Component TAS format") on its Central
Accounting Reporting System (CARS) page, and documents the Federal Account
Symbols and Titles (FAST) Book as three parts -- receipt accounts (Part I),
appropriation and other fund accounts (Part II), and foreign currency
accounts (Part III) -- each arranged within a documented set of fund groups,
on its "Description of Contents" page. Neither page publishes a machine
readable code list or a stable release identifier; each instead carries only
a "Last Updated" date, which this module treats as that page's edition.

This module captures those two pages and Treasury's official Part II and III
Excel workbook.  It parses every published workbook account row while
retaining Treasury's own TAS string as the account identifier.  Part I remains
a formatted PDF and is outside the workbook reader.  It never mints concept
identity for an account title: the rows are fiscal account metadata and the
identifier is Treasury's published value.

fiscal.treasury.gov pages embed per-request Akamai/Boomerang analytics
tokens (request IDs, timestamps) directly in the page body, so two live
captures of the identical logical page do not share one stable digest. This
module therefore pins each capture's own observed digest for cache and
local-file integrity, the same way every other registry acquisition module
does, rather than asserting one eternal "the" official page hash.

Live retrieval is provider-independent. Callers inject a fetcher or provide
an already captured local file. Importing this module never opens a network
connection.
"""

from __future__ import annotations

import hashlib
import os
import re
import tempfile
from collections.abc import Mapping
from dataclasses import dataclass
from datetime import UTC, date, datetime
from html.parser import HTMLParser
from io import BytesIO
from pathlib import Path
from typing import Literal, Protocol, cast
from urllib.parse import urlsplit

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from refspec.registry.infrastructure.controlled_identifier import ControlledIdentifier

TREASURY_PUBLISHER = "Bureau of the Fiscal Service, U.S. Department of the Treasury"
TREASURY_IDENTIFIER_AUTHORITY_URI = "https://fiscal.treasury.gov/accounting/"

ResourceName = Literal["tasComponentFormat", "fastBookDescriptionOfContents"]
AcquisitionMode = Literal["cache", "local", "fetcher"]
FASTBookPart = Literal["I", "II", "III"]
FundGroup = Literal["general", "special", "trust", "revolving", "deposit", "foreignCurrency"]

_DIGEST = re.compile(r"^sha256:([0-9a-f]{64})$")
_LAST_UPDATED = re.compile(r"Last Updated:\s*([A-Z][a-z]+ \d{1,2}, \d{4})")
_AID_PATTERN = re.compile(r"^\d{3}$")
_MAIN_PATTERN = re.compile(r"^\d{4}$")
_SUB_PATTERN = re.compile(r"^\d{3}$")
_ATA_PATTERN = re.compile(r"^\d{3}$")
_SP_PATTERN = re.compile(r"^\d{2}$")
_POA_PATTERN = re.compile(r"^\d{4}$")
_AVAILABILITY_TYPE_PATTERN = re.compile(r"^[A-Z]$")
_DEFAULT_SUB_ACCOUNT = "000"
_PUBLISHED_TAS_PATTERN = re.compile(r"^(?P<aid>\d{3})(?:X| )(?P<main>\d{4}(?:\.\d{3})?) ?$")
_FAST_BOOK_WORKBOOK_HEADERS: Mapping[str, tuple[str, ...]] = {
    "Part II": (
        "AID",
        "Main",
        "X-YEAR",
        "TAS",
        "Agency",
        "Title",
        "Legislation",
        "Fund Type",
        "Independent Agencies",
        "Last update",
    ),
    "Part III": (
        "AID",
        "Main",
        "X-YEAR",
        "TAS",
        "Agency",
        "Title",
        "Fund Type",
        "Independent Agencies",
        "Last update",
    ),
}
_FAST_BOOK_WORKBOOK_SHEETS = ("Intro Part II", "Part II", "Intro Part III", "Part III", "Changes")

# Treasury's own "Component TAS-BETC" flyer (a Bureau of the Fiscal Service
# document distributed by GPO) states the component sizes in order -- SP (2),
# ATA (3), AID (3), BPOA (4), EPOA (4), A (1), MAIN (4), SUB (3) -- which the
# width patterns above implement exactly. The PDF is pinned as reference-only
# provenance (fixture ``component-tas-betc-flyer.pdf``) and never parsed at
# runtime; its size row reads "Size: (2) (3) (3) (4) (4) (1) (4) (3)".
TAS_COMPONENT_SIZE_AUTHORITY_URL = (
    "https://www.gpo.gov/docs/default-source/guides-and-instructions/pdf/2-component-tas-betc-flyer.pdf"
)
TAS_COMPONENT_SIZE_AUTHORITY_SHA256 = "sha256:7a43d33c291a9ab233ed76237dec3cbfb28ebd10169183328077227f6c0cf2ea"
TAS_COMPONENT_SIZE_AUTHORITY_BYTE_LENGTH = 537_393
TAS_COMPONENT_SIZE_AUTHORITY_RETRIEVED_AT = "2026-08-03T23:31:00Z"

# The exact fund-group phrases the FAST Book "Description of Contents" page
# uses for Part I and Part II. Part III's paragraph names foreign currency
# accounts without enumerating a fund-group list, so RefSpec treats Part III
# as its own single fund group rather than inferring one.
PART_FUND_GROUPS: Mapping[FASTBookPart, tuple[FundGroup, ...]] = {
    "I": ("general", "special", "trust"),
    "II": ("general", "revolving", "special", "deposit", "trust"),
    "III": ("foreignCurrency",),
}

# Fund groups that the Description of Contents page says carry a citation to
# the United States Code or United States Statutes at Large. "general" fund
# accounts are not listed for either part.
_CITED_FUND_GROUPS: Mapping[FASTBookPart, frozenset[FundGroup]] = {
    "I": frozenset({"special", "trust"}),
    "II": frozenset({"revolving", "special", "deposit", "trust"}),
    "III": frozenset(),
}

TREASURY_TAS_FAST_BOOK_GAPS = (
    (
        "The Component TAS format page names its eight component fields (SP, ATA, "
        "AID, BPOA, EPOA, A, MAIN, SUB) but does not publish their exact character "
        'widths on that page; the widths applied here are publisher-stated in Treasury\'s own "Component '
        'TAS-BETC" flyer (pinned at TAS_COMPONENT_SIZE_AUTHORITY_SHA256, size row "Size: (2) (3) (3) (4) '
        '(4) (1) (4) (3)"), which confirms the OMB Circular A-11 / Governmentwide Spending Data Model '
        "convention this module started from. The flyer is reference-only provenance, not parsed at runtime."
    ),
    (
        "The official FAST Book publishes Part II and III account rows as an Excel "
        "workbook, which RefSpec pins and parses in full. Part I receipt accounts remain "
        "a formatted PDF and are not included in the workbook account collection."
    ),
    (
        "fiscal.treasury.gov pages embed per-request Akamai/Boomerang analytics "
        "tokens and timestamps directly in the page body, so two live captures of "
        "the identical logical page do not share one byte-stable digest."
    ),
)


class TreasuryTASFastBookError(ValueError):
    """Base class for Treasury TAS/FAST Book capture failures."""


class TreasuryAcquisitionError(TreasuryTASFastBookError):
    """Exact official source bytes could not be acquired safely."""


class TreasurySourceDriftError(TreasuryTASFastBookError):
    """A captured Treasury page no longer matches the reviewed structure."""


class TASComponentError(TreasuryTASFastBookError):
    """A Treasury Account Symbol component record is malformed or inconsistent."""


class FASTBookRecordError(TreasuryTASFastBookError):
    """A FAST Book account-title record is malformed or inconsistent with its edition."""


@dataclass(frozen=True, slots=True)
class TreasuryPageSource:
    """One official fiscal.treasury.gov "Description of Contents"-style page."""

    resource_name: ResourceName
    source_url: str
    filename: str
    structural_markers: tuple[str, ...]

    def __post_init__(self) -> None:
        parsed = urlsplit(self.source_url)
        if parsed.scheme != "https" or parsed.hostname != "fiscal.treasury.gov":
            raise TreasuryAcquisitionError("source_url must be an official HTTPS fiscal.treasury.gov URL")
        if parsed.username is not None or parsed.password is not None:
            raise TreasuryAcquisitionError("source_url must not contain credentials")
        if not self.filename or Path(self.filename).name != self.filename:
            raise TreasuryAcquisitionError("filename must be one plain path component")
        if not self.structural_markers:
            raise TreasuryAcquisitionError("structural_markers must not be empty")


TAS_COMPONENT_FORMAT_SOURCE = TreasuryPageSource(
    resource_name="tasComponentFormat",
    source_url=(
        "https://fiscal.treasury.gov/accounting/central-accounting-reporting-system-cars/"
        "treasury-account-symbol-reporting"
    ),
    filename="treasury-account-symbol-reporting.html",
    structural_markers=(
        "sub-level prefix (SP)",
        "allocation transfer identifier (ATA)",
        "agency identifier (AID)",
        "beginning period of availability (BPOA)",
        "ending period of availability (EPOA)",
        "availability type (A)",
        "main account (main)",
        "sub-account code (SUB)",
    ),
)
FAST_BOOK_DESCRIPTION_SOURCE = TreasuryPageSource(
    resource_name="fastBookDescriptionOfContents",
    source_url="https://fiscal.treasury.gov/accounting/fast-book/description-of-contents",
    filename="fast-book-description-of-contents.html",
    structural_markers=(
        "Part I contains receipt accounts arranged numerically within each fund group",
        "general, special and trust",
        "Part II contains appropriation and other fund accounts for each agency",
        "general, revolving, special, deposit and trust",
        "Part III contains foreign currency accounts",
    ),
)


@dataclass(frozen=True, slots=True)
class TreasuryPageSnapshotPin:
    """Exact identity of one Treasury page capture."""

    source: TreasuryPageSource
    retrieved_at: str
    expected_sha256: str
    expected_byte_length: int

    def __post_init__(self) -> None:
        if _DIGEST.fullmatch(self.expected_sha256) is None:
            raise TreasuryAcquisitionError("expected_sha256 must be a lowercase sha256:<64 hex> digest")
        if self.expected_byte_length <= 0:
            raise TreasuryAcquisitionError("expected_byte_length must be positive")
        if not self.retrieved_at.strip():
            raise TreasuryAcquisitionError("retrieved_at must not be empty")


# Real captures made 2026-08-03. Pinned here for cache/local-file integrity
# checks, not as a claim that a future live fetch will reproduce this exact
# digest (see the Akamai/Boomerang gap recorded above).
TAS_COMPONENT_FORMAT_2026_08_03 = TreasuryPageSnapshotPin(
    source=TAS_COMPONENT_FORMAT_SOURCE,
    retrieved_at="2026-08-03T19:17:15Z",
    expected_sha256="sha256:fbd8c6794fdf10d4e1b28ece79af5c15352eb25d292069e0238c3c7513f4675d",
    expected_byte_length=112_908,
)
FAST_BOOK_DESCRIPTION_2026_08_03 = TreasuryPageSnapshotPin(
    source=FAST_BOOK_DESCRIPTION_SOURCE,
    retrieved_at="2026-08-03T19:17:43Z",
    expected_sha256="sha256:91525d80cc4bd6e8ab08075ad630b484b0f691c08516a36151589ddbd57c2a36",
    expected_byte_length=110_043,
)


@dataclass(frozen=True, slots=True)
class FASTBookWorkbookPin:
    """Exact identity and reviewed edition of an official TFX workbook."""

    source_url: str
    filename: str
    retrieved_at: str
    edition: str
    expected_sha256: str
    expected_byte_length: int
    expected_modified_at: str
    expected_part_ii_rows: int
    expected_part_iii_rows: int
    expected_change_rows: int

    def __post_init__(self) -> None:
        parsed = urlsplit(self.source_url)
        if parsed.scheme != "https" or parsed.hostname != "tfx.treasury.gov":
            raise TreasuryAcquisitionError("FAST Book workbook must use official HTTPS tfx.treasury.gov")
        if parsed.username is not None or parsed.password is not None:
            raise TreasuryAcquisitionError("FAST Book workbook URL must not contain credentials")
        if not self.filename.endswith(".xlsx") or Path(self.filename).name != self.filename:
            raise TreasuryAcquisitionError("FAST Book workbook filename must be one .xlsx path component")
        if _DIGEST.fullmatch(self.expected_sha256) is None:
            raise TreasuryAcquisitionError("expected_sha256 must be a lowercase sha256:<64 hex> digest")
        if self.expected_byte_length <= 0:
            raise TreasuryAcquisitionError("expected_byte_length must be positive")
        if re.fullmatch(r"\d{4}-\d{2}", self.edition) is None:
            raise TreasuryAcquisitionError("FAST Book workbook edition must be YYYY-MM")
        try:
            datetime.fromisoformat(self.expected_modified_at)
        except ValueError as error:
            raise TreasuryAcquisitionError("expected_modified_at must be an ISO 8601 timestamp") from error
        if min(self.expected_part_ii_rows, self.expected_part_iii_rows, self.expected_change_rows) <= 0:
            raise TreasuryAcquisitionError("FAST Book expected workbook row counts must be positive")


FAST_BOOK_PART_II_III_SOURCE_URL = "https://tfx.treasury.gov/media/60111/download?inline="
FAST_BOOK_PART_II_III_2026_07_31 = FASTBookWorkbookPin(
    source_url=FAST_BOOK_PART_II_III_SOURCE_URL,
    filename="fast-book-part-ii-iii-2026-07-31.xlsx",
    retrieved_at="2026-08-04T04:36:30Z",
    edition="2026-07",
    expected_sha256="sha256:0e40902a2e4bfee7439fbe24d90fd9ff39fad859b4ba432725256866b06cb461",
    expected_byte_length=420_508,
    expected_modified_at="2026-07-30T19:11:58",
    expected_part_ii_rows=3_442,
    expected_part_iii_rows=140,
    expected_change_rows=1_159,
)


@dataclass(frozen=True, slots=True)
class FetchedTreasuryPage:
    """Provider-independent response returned by an injected fetcher."""

    body: bytes
    status_code: int
    content_type: str
    resolved_url: str


class TreasuryPageFetcher(Protocol):
    """Small transport boundary for official fiscal.treasury.gov pages."""

    def fetch(self, source_url: str, *, timeout_seconds: float) -> FetchedTreasuryPage:
        """Fetch one page while preserving its exact body bytes."""


@dataclass(frozen=True, slots=True)
class AcquiredTreasuryPage:
    """One verified Treasury page in the content-addressed store."""

    pin: TreasuryPageSnapshotPin
    path: Path
    sha256: str
    byte_length: int
    source_url: str
    resolved_url: str | None
    content_type: str
    acquisition_mode: AcquisitionMode
    cache_hit: bool
    local_source_path: Path | None


def sha256_digest(payload: bytes) -> str:
    """Return the canonical RefSpec SHA-256 spelling."""

    return "sha256:" + hashlib.sha256(payload).hexdigest()


def _validate_resolved_url(value: str) -> None:
    parsed = urlsplit(value)
    if parsed.scheme != "https" or parsed.hostname != "fiscal.treasury.gov":
        raise TreasuryAcquisitionError("fetcher resolved_url must remain on official HTTPS fiscal.treasury.gov")
    if parsed.username is not None or parsed.password is not None:
        raise TreasuryAcquisitionError("fetcher resolved_url must not contain credentials")


def _verify_payload(payload: bytes, pin: TreasuryPageSnapshotPin, *, location: str) -> tuple[str, int]:
    byte_length = len(payload)
    if byte_length != pin.expected_byte_length:
        raise TreasurySourceDriftError(
            f"{location} byte length drift: expected {pin.expected_byte_length}, got {byte_length}"
        )
    actual_sha256 = sha256_digest(payload)
    if actual_sha256 != pin.expected_sha256:
        raise TreasurySourceDriftError(f"{location} digest drift: expected {pin.expected_sha256}, got {actual_sha256}")
    lowered = payload[:64_000].lower()
    if b"<html" not in lowered and b"<!doctype html" not in lowered:
        raise TreasurySourceDriftError(f"{location} is not an HTML document")
    return actual_sha256, byte_length


def _verify_existing(path: Path, pin: TreasuryPageSnapshotPin) -> AcquiredTreasuryPage:
    if path.is_symlink() or not path.is_file():
        raise TreasuryAcquisitionError(f"content-addressed target is not a regular file: {path}")
    actual_sha256, byte_length = _verify_payload(path.read_bytes(), pin, location="cached Treasury page")
    return AcquiredTreasuryPage(
        pin=pin,
        path=path,
        sha256=actual_sha256,
        byte_length=byte_length,
        source_url=pin.source.source_url,
        resolved_url=None,
        content_type="text/html",
        acquisition_mode="cache",
        cache_hit=True,
        local_source_path=None,
    )


def _publish_payload(
    payload: bytes,
    pin: TreasuryPageSnapshotPin,
    final_path: Path,
    *,
    content_type: str,
    acquisition_mode: Literal["local", "fetcher"],
    resolved_url: str | None,
    local_source_path: Path | None,
) -> AcquiredTreasuryPage:
    actual_sha256, byte_length = _verify_payload(payload, pin, location=f"{acquisition_mode} Treasury page")
    final_path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(prefix=".acquire-", suffix=".tmp", dir=final_path.parent)
    temporary_path = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as output:
            descriptor = -1
            output.write(payload)
            output.flush()
            os.fsync(output.fileno())
        try:
            os.link(temporary_path, final_path)
        except FileExistsError:
            return _verify_existing(final_path, pin)
        return AcquiredTreasuryPage(
            pin=pin,
            path=final_path,
            sha256=actual_sha256,
            byte_length=byte_length,
            source_url=pin.source.source_url,
            resolved_url=resolved_url,
            content_type=content_type,
            acquisition_mode=acquisition_mode,
            cache_hit=False,
            local_source_path=local_source_path,
        )
    finally:
        if descriptor >= 0:
            os.close(descriptor)
        temporary_path.unlink(missing_ok=True)


def acquire_treasury_page(
    pin: TreasuryPageSnapshotPin,
    store_dir: Path,
    *,
    source_path: Path | None = None,
    fetcher: TreasuryPageFetcher | None = None,
    timeout_seconds: float = 30.0,
) -> AcquiredTreasuryPage:
    """Acquire one exact page through a provider-neutral boundary."""

    if timeout_seconds <= 0:
        raise TreasuryAcquisitionError("timeout_seconds must be positive")
    if source_path is not None and fetcher is not None:
        raise TreasuryAcquisitionError("provide source_path or fetcher, not both")
    digest_hex = cast(re.Match[str], _DIGEST.fullmatch(pin.expected_sha256)).group(1)
    final_path = Path(store_dir) / "sha256" / digest_hex / pin.source.filename
    if final_path.exists() or final_path.is_symlink():
        return _verify_existing(final_path, pin)

    if source_path is not None:
        local_path = Path(source_path)
        if local_path.is_symlink() or not local_path.is_file():
            raise TreasuryAcquisitionError(f"local Treasury source is not a regular file: {local_path}")
        return _publish_payload(
            local_path.read_bytes(),
            pin,
            final_path,
            content_type="text/html",
            acquisition_mode="local",
            resolved_url=None,
            local_source_path=local_path.resolve(),
        )

    if fetcher is None:
        raise TreasuryAcquisitionError("Treasury page is not cached; provide source_path or an injected fetcher")
    fetched = fetcher.fetch(pin.source.source_url, timeout_seconds=timeout_seconds)
    if fetched.status_code != 200:
        raise TreasuryAcquisitionError(f"could not acquire {pin.source.source_url}: HTTP {fetched.status_code}")
    _validate_resolved_url(fetched.resolved_url)
    media_type = fetched.content_type.partition(";")[0].strip().lower()
    if media_type not in {"text/html", "application/xhtml+xml"}:
        raise TreasurySourceDriftError(f"Treasury page content type drifted to {fetched.content_type!r}")
    return _publish_payload(
        fetched.body,
        pin,
        final_path,
        content_type=fetched.content_type,
        acquisition_mode="fetcher",
        resolved_url=fetched.resolved_url,
        local_source_path=None,
    )


class _TextCollector(HTMLParser):
    """Collect visible page text, dropping script/style contents."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self._skip_depth = 0
        self.chunks: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        del attrs
        if tag in {"script", "style"}:
            self._skip_depth += 1

    def handle_endtag(self, tag: str) -> None:
        if tag in {"script", "style"} and self._skip_depth:
            self._skip_depth -= 1

    def handle_data(self, data: str) -> None:
        if not self._skip_depth:
            self.chunks.append(data)


def _extract_text(payload: bytes) -> str:
    try:
        decoded = payload.decode("utf-8")
    except UnicodeDecodeError as error:
        raise TreasurySourceDriftError("Treasury page is not UTF-8") from error
    collector = _TextCollector()
    try:
        collector.feed(decoded)
        collector.close()
    except Exception as error:
        raise TreasurySourceDriftError("Treasury page is malformed HTML") from error
    return " ".join("".join(collector.chunks).split())


def _extract_edition_date(text: str, *, location: str) -> str:
    match = _LAST_UPDATED.search(text)
    if match is None:
        raise TreasurySourceDriftError(f"{location} is missing its 'Last Updated:' edition marker")
    try:
        parsed = datetime.strptime(match.group(1), "%B %d, %Y").replace(tzinfo=UTC).date()
    except ValueError as error:
        raise TreasurySourceDriftError(f"{location} 'Last Updated:' date is unparseable") from error
    return parsed.isoformat()


def _read_acquired_payload(page: AcquiredTreasuryPage) -> bytes:
    payload = page.path.read_bytes()
    _verify_payload(payload, page.pin, location="parsed Treasury page")
    return payload


@dataclass(frozen=True, slots=True)
class ParsedTASComponentFormat:
    """The Component TAS field list and edition read from one exact capture."""

    source_sha256: str
    source_byte_length: int
    retrieved_at: str
    edition_date: str
    component_field_labels: tuple[str, ...]
    gaps: tuple[str, ...]


def parse_tas_component_page(page: AcquiredTreasuryPage) -> ParsedTASComponentFormat:
    """Parse the Component TAS format page's field list and edition date."""

    payload = _read_acquired_payload(page)
    text = _extract_text(payload)
    for marker in TAS_COMPONENT_FORMAT_SOURCE.structural_markers:
        if marker not in text:
            raise TreasurySourceDriftError(f"missing expected component marker {marker!r}")
    edition_date = _extract_edition_date(text, location="Component TAS format page")
    return ParsedTASComponentFormat(
        source_sha256=page.sha256,
        source_byte_length=page.byte_length,
        retrieved_at=page.pin.retrieved_at,
        edition_date=edition_date,
        component_field_labels=TAS_COMPONENT_FORMAT_SOURCE.structural_markers,
        gaps=TREASURY_TAS_FAST_BOOK_GAPS,
    )


@dataclass(frozen=True, slots=True)
class ParsedFASTBookDescription:
    """The FAST Book part/fund-group structure and edition read from one capture."""

    source_sha256: str
    source_byte_length: int
    retrieved_at: str
    edition_date: str
    part_fund_groups: Mapping[FASTBookPart, tuple[FundGroup, ...]]
    gaps: tuple[str, ...]


def parse_fast_book_description_page(page: AcquiredTreasuryPage) -> ParsedFASTBookDescription:
    """Parse the FAST Book Description of Contents page's structure and edition."""

    payload = _read_acquired_payload(page)
    text = _extract_text(payload)
    for marker in FAST_BOOK_DESCRIPTION_SOURCE.structural_markers:
        if marker not in text:
            raise TreasurySourceDriftError(f"missing expected FAST Book marker {marker!r}")
    edition_date = _extract_edition_date(text, location="FAST Book Description of Contents page")
    return ParsedFASTBookDescription(
        source_sha256=page.sha256,
        source_byte_length=page.byte_length,
        retrieved_at=page.pin.retrieved_at,
        edition_date=edition_date,
        part_fund_groups=PART_FUND_GROUPS,
        gaps=TREASURY_TAS_FAST_BOOK_GAPS,
    )


# ---------------------------------------------------------------------------
# FAST Book Part II and III official workbook
# ---------------------------------------------------------------------------


@dataclass(frozen=True, slots=True)
class FASTBookPublishedAccount:
    """One account row exactly identified by Treasury's published TAS value."""

    part: Literal["II", "III"]
    treasury_account_symbol: str
    agency_identifier: str
    main_account: str
    agency_name: str
    account_title: str
    legislation: str | None
    fund_type: str
    independent_agency_identifier: str | None
    last_updated: str | None

    def __post_init__(self) -> None:
        match = _PUBLISHED_TAS_PATTERN.fullmatch(self.treasury_account_symbol)
        if match is None:
            raise FASTBookRecordError(f"published TAS has an unexpected shape: {self.treasury_account_symbol!r}")
        if self.part not in {"II", "III"}:
            raise FASTBookRecordError("published FAST Book part must be II or III")
        if self.agency_identifier != match.group("aid"):
            raise FASTBookRecordError("agency_identifier must match the published TAS")
        if self.main_account != match.group("main"):
            raise FASTBookRecordError("main_account must match the published TAS")
        for name, value in (
            ("agency_name", self.agency_name),
            ("account_title", self.account_title),
            ("fund_type", self.fund_type),
        ):
            if not value or value != value.strip():
                raise FASTBookRecordError(f"{name} must be non-empty normalized text")
        if self.legislation is not None and (not self.legislation or self.legislation != self.legislation.strip()):
            raise FASTBookRecordError("legislation must be normalized text or null")
        if (
            self.independent_agency_identifier is not None
            and _AID_PATTERN.fullmatch(self.independent_agency_identifier) is None
        ):
            raise FASTBookRecordError("independent_agency_identifier must be exactly 3 digits or null")
        if self.last_updated is not None:
            try:
                date.fromisoformat(self.last_updated)
            except ValueError as error:
                raise FASTBookRecordError("last_updated must be an ISO 8601 date or null") from error


@dataclass(frozen=True, slots=True)
class ParsedFASTBookWorkbook:
    """All official Part II and III account rows from one exact workbook."""

    source_url: str
    source_sha256: str
    source_byte_length: int
    retrieved_at: str
    edition: str
    workbook_modified_at: str
    part_ii_row_count: int
    part_iii_row_count: int
    change_row_count: int
    accounts: tuple[FASTBookPublishedAccount, ...]
    publisher_anomalies: tuple[str, ...]


def _normalized_workbook_text(value: object, *, sheet: str, row_number: int, field: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise TreasurySourceDriftError(f"{sheet} row {row_number} {field} must be non-empty text")
    return value.strip()


def _optional_workbook_text(value: object, *, sheet: str, row_number: int, field: str) -> str | None:
    if value is None:
        return None
    if not isinstance(value, str) or not value.strip():
        raise TreasurySourceDriftError(f"{sheet} row {row_number} {field} must be text or empty")
    return value.strip()


def _independent_agency_identifier(value: object, *, sheet: str, row_number: int) -> str | None:
    if value is None:
        return None
    if isinstance(value, int) and 0 <= value <= 999:
        return f"{value:03d}"
    if isinstance(value, str) and value.strip().isdigit() and len(value.strip()) <= 3:
        return value.strip().zfill(3)
    raise TreasurySourceDriftError(f"{sheet} row {row_number} Independent Agencies has an unexpected value")


def _workbook_date(value: object, *, sheet: str, row_number: int) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raise TreasurySourceDriftError(f"{sheet} row {row_number} Last update is not an Excel date")


def _workbook_main(value: object) -> str | None:
    if isinstance(value, int) and 0 <= value <= 9999:
        return f"{value:04d}"
    if isinstance(value, float) and 0 <= value < 10_000:
        return f"{value:08.3f}"
    return None


def _parse_published_account_row(
    row: tuple[object, ...],
    *,
    part: Literal["II", "III"],
    sheet: str,
    row_number: int,
    anomalies: list[str],
) -> FASTBookPublishedAccount:
    tas_value = _normalized_workbook_text(row[3], sheet=sheet, row_number=row_number, field="TAS")
    tas_match = _PUBLISHED_TAS_PATTERN.fullmatch(tas_value)
    if tas_match is None:
        raise TreasurySourceDriftError(f"{sheet} row {row_number} TAS has an unexpected shape: {tas_value!r}")

    published_aid = row[0]
    normalized_aid = (
        f"{published_aid:03d}"
        if isinstance(published_aid, int) and 0 <= published_aid <= 999
        else str(published_aid).strip().zfill(3)
    )
    if normalized_aid != tas_match.group("aid"):
        anomalies.append(
            f"{sheet} row {row_number}: AID cell {published_aid!r} does not match TAS {tas_value!r}; TAS retained"
        )

    workbook_main = _workbook_main(row[1])
    tas_main = tas_match.group("main")
    comparable_tas_main = tas_main if workbook_main is not None and "." in workbook_main else tas_main.partition(".")[0]
    if workbook_main != comparable_tas_main:
        anomalies.append(
            f"{sheet} row {row_number}: Main cell {row[1]!r} does not match TAS {tas_value!r}; TAS retained"
        )

    duration_cell = row[2]
    duration_value = " " if duration_cell is None else str(duration_cell).strip() or " "
    if duration_value != tas_value[3]:
        anomalies.append(
            f"{sheet} row {row_number}: X-YEAR cell {duration_cell!r} does not match TAS {tas_value!r}; TAS retained"
        )

    if part == "II":
        legislation = _optional_workbook_text(row[6], sheet=sheet, row_number=row_number, field="Legislation")
        fund_type_cell, independent_cell, update_cell = row[7], row[8], row[9]
    else:
        legislation = None
        fund_type_cell, independent_cell, update_cell = row[6], row[7], row[8]

    last_updated = _workbook_date(update_cell, sheet=sheet, row_number=row_number)
    if last_updated is None:
        anomalies.append(f"{sheet} row {row_number}: Last update is empty in the publisher workbook")

    return FASTBookPublishedAccount(
        part=part,
        treasury_account_symbol=tas_value,
        agency_identifier=tas_match.group("aid"),
        main_account=tas_match.group("main"),
        agency_name=_normalized_workbook_text(row[4], sheet=sheet, row_number=row_number, field="Agency"),
        account_title=_normalized_workbook_text(row[5], sheet=sheet, row_number=row_number, field="Title"),
        legislation=legislation,
        fund_type=_normalized_workbook_text(
            fund_type_cell,
            sheet=sheet,
            row_number=row_number,
            field="Fund Type",
        ),
        independent_agency_identifier=_independent_agency_identifier(
            independent_cell,
            sheet=sheet,
            row_number=row_number,
        ),
        last_updated=last_updated,
    )


def parse_fast_book_workbook(
    source_path: Path,
    *,
    pin: FASTBookWorkbookPin,
) -> ParsedFASTBookWorkbook:
    """Parse every Part II and III row from one exact official workbook.

    The workbook's published ``TAS`` column is authoritative.  The separate
    convenience columns have a small number of known publisher defects; those
    are reported in ``publisher_anomalies`` and never used to rewrite the TAS.
    """

    path = Path(source_path)
    if path.is_symlink() or not path.is_file():
        raise TreasuryAcquisitionError(f"FAST Book workbook is not a regular file: {path}")
    payload = path.read_bytes()
    if len(payload) != pin.expected_byte_length:
        raise TreasurySourceDriftError(
            f"FAST Book workbook byte length drift: expected {pin.expected_byte_length}, got {len(payload)}"
        )
    digest = sha256_digest(payload)
    if digest != pin.expected_sha256:
        raise TreasurySourceDriftError(
            f"FAST Book workbook digest drift: expected {pin.expected_sha256}, got {digest}"
        )
    if not payload.startswith(b"PK"):
        raise TreasurySourceDriftError("FAST Book workbook is not an XLSX ZIP document")

    try:
        workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    except (InvalidFileException, OSError, ValueError) as error:
        raise TreasurySourceDriftError("FAST Book workbook is not readable XLSX") from error

    if tuple(workbook.sheetnames) != _FAST_BOOK_WORKBOOK_SHEETS:
        raise TreasurySourceDriftError(
            f"FAST Book workbook sheets drifted: expected {_FAST_BOOK_WORKBOOK_SHEETS!r}, "
            f"got {tuple(workbook.sheetnames)!r}"
        )
    modified = workbook.properties.modified
    if modified is None or modified.isoformat() != pin.expected_modified_at:
        actual_modified = None if modified is None else modified.isoformat()
        raise TreasurySourceDriftError(
            f"FAST Book workbook modified timestamp drifted: expected {pin.expected_modified_at}, "
            f"got {actual_modified}"
        )
    if not pin.expected_modified_at.startswith(pin.edition):
        raise TreasurySourceDriftError("FAST Book pin edition does not match the workbook modified month")

    accounts: list[FASTBookPublishedAccount] = []
    anomalies: list[str] = []
    part_counts: dict[str, int] = {}
    seen_tas: dict[str, tuple[str, int]] = {}
    for sheet, part in (("Part II", "II"), ("Part III", "III")):
        worksheet = workbook[sheet]
        header = tuple(cell.value for cell in next(worksheet.iter_rows(min_row=2, max_row=2)))
        if header != _FAST_BOOK_WORKBOOK_HEADERS[sheet]:
            raise TreasurySourceDriftError(
                f"{sheet} headers drifted: expected {_FAST_BOOK_WORKBOOK_HEADERS[sheet]!r}, got {header!r}"
            )
        count = 0
        for row_number, cells in enumerate(worksheet.iter_rows(min_row=3, values_only=True), start=3):
            row = tuple(cells)
            if not any(value is not None for value in row):
                continue
            account = _parse_published_account_row(
                row,
                part=cast(Literal["II", "III"], part),
                sheet=sheet,
                row_number=row_number,
                anomalies=anomalies,
            )
            previous = seen_tas.get(account.treasury_account_symbol)
            if previous is not None:
                anomalies.append(
                    f"{sheet} row {row_number}: TAS {account.treasury_account_symbol!r} duplicates "
                    f"{previous[0]} row {previous[1]}; both publisher rows retained"
                )
            else:
                seen_tas[account.treasury_account_symbol] = (sheet, row_number)
            accounts.append(account)
            count += 1
        part_counts[part] = count

    changes = workbook["Changes"]
    change_header = tuple(cell.value for cell in next(changes.iter_rows(min_row=1, max_row=1)))
    expected_change_header = (*_FAST_BOOK_WORKBOOK_HEADERS["Part II"], "Action", "Comments")
    if change_header != expected_change_header:
        raise TreasurySourceDriftError("FAST Book Changes headers drifted")
    change_count = sum(
        1 for row in changes.iter_rows(min_row=2, values_only=True) if any(value is not None for value in row)
    )

    actual_counts = (part_counts["II"], part_counts["III"], change_count)
    expected_counts = (pin.expected_part_ii_rows, pin.expected_part_iii_rows, pin.expected_change_rows)
    if actual_counts != expected_counts:
        raise TreasurySourceDriftError(
            f"FAST Book workbook row counts drifted: expected {expected_counts!r}, got {actual_counts!r}"
        )

    return ParsedFASTBookWorkbook(
        source_url=pin.source_url,
        source_sha256=digest,
        source_byte_length=len(payload),
        retrieved_at=pin.retrieved_at,
        edition=pin.edition,
        workbook_modified_at=pin.expected_modified_at,
        part_ii_row_count=part_counts["II"],
        part_iii_row_count=part_counts["III"],
        change_row_count=change_count,
        accounts=tuple(accounts),
        publisher_anomalies=tuple(anomalies),
    )


def published_fast_book_identifier(
    record: FASTBookPublishedAccount,
    *,
    observed_at: str | None,
    source_digest: str | None,
) -> ControlledIdentifier:
    """Return Treasury's exact published TAS as a source-anchored identifier."""

    return ControlledIdentifier(
        value=record.treasury_account_symbol,
        kind="treasuryAccountSymbol",
        authority_uri=TREASURY_IDENTIFIER_AUTHORITY_URI,
        source_uri=FAST_BOOK_PART_II_III_SOURCE_URL,
        observed_at=observed_at,
        effective_at=record.last_updated,
        source_digest=source_digest,
    )


# ---------------------------------------------------------------------------
# Treasury Account Symbol component identifier shape
# ---------------------------------------------------------------------------

_TAS_RECORD_FIELDS = ("SP", "ATA", "AID", "BPOA", "EPOA", "A", "MAIN", "SUB")
_TAS_REQUIRED_FIELDS = frozenset({"AID", "MAIN"})


@dataclass(frozen=True, slots=True)
class TASComponents:
    """One Treasury Account Symbol, retained as its eight named components.

    Field widths follow the published OMB Circular A-11 / Governmentwide
    Spending Data Model Component TAS convention; see
    ``TREASURY_TAS_FAST_BOOK_GAPS`` for why the captured Treasury page itself
    does not confirm them.
    """

    sub_level_prefix: str | None
    allocation_transfer_agency: str | None
    agency_identifier: str
    beginning_period_of_availability: str | None
    ending_period_of_availability: str | None
    availability_type_code: str | None
    main_account: str
    sub_account: str

    def __post_init__(self) -> None:
        if self.sub_level_prefix is not None and _SP_PATTERN.fullmatch(self.sub_level_prefix) is None:
            raise TASComponentError("sub_level_prefix must be exactly 2 digits")
        if (
            self.allocation_transfer_agency is not None
            and _ATA_PATTERN.fullmatch(self.allocation_transfer_agency) is None
        ):
            raise TASComponentError("allocation_transfer_agency must be exactly 3 digits")
        if _AID_PATTERN.fullmatch(self.agency_identifier) is None:
            raise TASComponentError("agency_identifier must be exactly 3 digits")
        if (
            self.beginning_period_of_availability is not None
            and _POA_PATTERN.fullmatch(self.beginning_period_of_availability) is None
        ):
            raise TASComponentError("beginning_period_of_availability must be exactly 4 digits")
        if (
            self.ending_period_of_availability is not None
            and _POA_PATTERN.fullmatch(self.ending_period_of_availability) is None
        ):
            raise TASComponentError("ending_period_of_availability must be exactly 4 digits")
        if (self.beginning_period_of_availability is None) != (self.ending_period_of_availability is None):
            raise TASComponentError(
                "beginning_period_of_availability and ending_period_of_availability "
                "must be given together or both left unset"
            )
        if (
            self.beginning_period_of_availability is not None
            and self.ending_period_of_availability is not None
            and self.beginning_period_of_availability > self.ending_period_of_availability
        ):
            raise TASComponentError("beginning_period_of_availability must not be after ending_period_of_availability")
        if (
            self.availability_type_code is not None
            and _AVAILABILITY_TYPE_PATTERN.fullmatch(self.availability_type_code) is None
        ):
            raise TASComponentError("availability_type_code must be exactly one uppercase letter")
        if _MAIN_PATTERN.fullmatch(self.main_account) is None:
            raise TASComponentError("main_account must be exactly 4 digits")
        if _SUB_PATTERN.fullmatch(self.sub_account) is None:
            raise TASComponentError("sub_account must be exactly 3 digits")


def parse_tas_components(record: Mapping[str, object]) -> TASComponents:
    """Parse one CARS-style Component TAS record into strict, typed fields."""

    unknown = set(record) - set(_TAS_RECORD_FIELDS)
    if unknown:
        raise TASComponentError(f"unknown Component TAS fields: {sorted(unknown)}")
    missing = _TAS_REQUIRED_FIELDS - set(record)
    if missing:
        raise TASComponentError(f"missing required Component TAS fields: {sorted(missing)}")

    def field(name: str, *, required: bool = False) -> str | None:
        value = record.get(name)
        if value is None:
            if required:
                raise TASComponentError(f"{name} must not be empty")
            return None
        if not isinstance(value, str) or not value.strip():
            raise TASComponentError(f"{name} must be non-empty text")
        return value

    sub_account = field("SUB") or _DEFAULT_SUB_ACCOUNT
    return TASComponents(
        sub_level_prefix=field("SP"),
        allocation_transfer_agency=field("ATA"),
        agency_identifier=cast(str, field("AID", required=True)),
        beginning_period_of_availability=field("BPOA"),
        ending_period_of_availability=field("EPOA"),
        availability_type_code=field("A"),
        main_account=cast(str, field("MAIN", required=True)),
        sub_account=sub_account,
    )


def _canonical_component(value: str | None) -> str:
    return value if value is not None else ""


def tas_identifier(
    components: TASComponents,
    *,
    observed_at: str | None,
    source_digest: str | None,
) -> ControlledIdentifier:
    """Build a capture-local, order-preserving identifier for one TAS.

    The returned value is a RefSpec-local dot-joined encoding of the eight
    Component TAS fields, in their documented order, with an absent field
    left blank between two dots. It is not a Treasury-published display
    string: Treasury's own systems render a TAS differently depending on the
    reporting context, and no single canonical string format was found.
    """

    value = ".".join(
        _canonical_component(component)
        for component in (
            components.sub_level_prefix,
            components.allocation_transfer_agency,
            components.agency_identifier,
            components.beginning_period_of_availability,
            components.ending_period_of_availability,
            components.availability_type_code,
            components.main_account,
            components.sub_account,
        )
    )
    return ControlledIdentifier(
        value=value,
        kind="treasuryAccountSymbolComponents",
        authority_uri=TREASURY_IDENTIFIER_AUTHORITY_URI,
        source_uri=TAS_COMPONENT_FORMAT_SOURCE.source_url,
        observed_at=observed_at,
        effective_at=None,
        source_digest=source_digest,
    )


def parse_tas_canonical_value(value: str) -> TASComponents:
    """Invert :func:`tas_identifier`'s canonical encoding back to components."""

    parts = value.split(".")
    if len(parts) != len(_TAS_RECORD_FIELDS):
        raise TASComponentError("canonical TAS value must have exactly eight dot-separated fields")
    record = {name: (part or None) for name, part in zip(_TAS_RECORD_FIELDS, parts, strict=True)}
    return parse_tas_components(record)


# ---------------------------------------------------------------------------
# FAST Book account-title record shape
# ---------------------------------------------------------------------------


@dataclass(frozen=True, slots=True)
class FASTBookAccountRecord:
    """One agency/account-title record, shaped like a FAST Book entry.

    This is an identifier-shape and code-value capture, not bulk FAST Book
    content: RefSpec does not fetch or store the FAST Book's PDF row data.
    """

    agency_identifier: str
    main_account: str
    account_title: str
    fast_book_part: FASTBookPart
    fund_group: FundGroup
    statutory_citation: str | None
    edition_date: str

    def __post_init__(self) -> None:
        if _AID_PATTERN.fullmatch(self.agency_identifier) is None:
            raise FASTBookRecordError("agency_identifier must be exactly 3 digits")
        if _MAIN_PATTERN.fullmatch(self.main_account) is None:
            raise FASTBookRecordError("main_account must be exactly 4 digits")
        if not self.account_title.strip():
            raise FASTBookRecordError("account_title must not be empty")
        if self.fast_book_part not in PART_FUND_GROUPS:
            raise FASTBookRecordError(f"fast_book_part must be one of {sorted(PART_FUND_GROUPS)}")
        if self.fund_group not in PART_FUND_GROUPS[self.fast_book_part]:
            raise FASTBookRecordError(
                f"fund_group {self.fund_group!r} is not a documented fund group for Part {self.fast_book_part}"
            )
        requires_citation = self.fund_group in _CITED_FUND_GROUPS[self.fast_book_part]
        if requires_citation and not (self.statutory_citation and self.statutory_citation.strip()):
            raise FASTBookRecordError(
                f"Part {self.fast_book_part} {self.fund_group} fund accounts requires a statutory citation"
            )
        if not requires_citation and self.statutory_citation is not None:
            raise FASTBookRecordError(
                f"Part {self.fast_book_part} {self.fund_group} fund general fund account "
                "must not carry a statutory citation"
            )
        if date.fromisoformat(self.edition_date) is None:
            raise FASTBookRecordError("edition_date must be an ISO 8601 date")  # pragma: no cover - defensive


def validate_fast_book_account_record(
    raw: Mapping[str, object],
    *,
    description: ParsedFASTBookDescription,
) -> FASTBookAccountRecord:
    """Validate one illustrative FAST Book-shaped record against its edition."""

    required = {"AID", "MAIN", "ACCOUNT_TITLE", "PART", "FUND_GROUP", "STATUTORY_CITATION"}
    if set(raw) != required:
        raise FASTBookRecordError(f"FAST Book record fields must be exactly {sorted(required)}")

    def text(name: str) -> str:
        value = raw[name]
        if not isinstance(value, str) or not value.strip():
            raise FASTBookRecordError(f"{name} must be non-empty text")
        return value

    citation = raw["STATUTORY_CITATION"]
    if citation is not None and (not isinstance(citation, str) or not citation.strip()):
        raise FASTBookRecordError("STATUTORY_CITATION must be non-empty text or null")

    return FASTBookAccountRecord(
        agency_identifier=text("AID"),
        main_account=text("MAIN"),
        account_title=text("ACCOUNT_TITLE"),
        fast_book_part=cast(FASTBookPart, text("PART")),
        fund_group=cast(FundGroup, text("FUND_GROUP")),
        statutory_citation=citation,
        edition_date=description.edition_date,
    )


def fast_book_identifier(
    record: FASTBookAccountRecord,
    *,
    observed_at: str | None,
    source_digest: str | None,
) -> ControlledIdentifier:
    """Build a capture-local identifier for one FAST Book-shaped account record."""

    return ControlledIdentifier(
        value=f"{record.agency_identifier}-{record.main_account}",
        kind="fastBookAccountIdentifier",
        authority_uri=TREASURY_IDENTIFIER_AUTHORITY_URI,
        source_uri=FAST_BOOK_DESCRIPTION_SOURCE.source_url,
        observed_at=observed_at,
        effective_at=None,
        source_digest=source_digest,
    )


# ---------------------------------------------------------------------------
# Deterministic fiscal edition
# ---------------------------------------------------------------------------


@dataclass(frozen=True, slots=True)
class TreasuryTASFastBookEdition:
    """The recorded edition of both captured pages, plus known gaps."""

    tas_component_format_edition: str
    fast_book_description_edition: str
    gaps: tuple[str, ...]


def assemble_treasury_tas_fast_book_edition(
    tas_page: ParsedTASComponentFormat,
    fast_book_page: ParsedFASTBookDescription,
) -> TreasuryTASFastBookEdition:
    """Combine both parsed pages' edition dates into one recorded edition."""

    return TreasuryTASFastBookEdition(
        tas_component_format_edition=tas_page.edition_date,
        fast_book_description_edition=fast_book_page.edition_date,
        gaps=TREASURY_TAS_FAST_BOOK_GAPS,
    )


__all__ = [
    "FAST_BOOK_DESCRIPTION_2026_08_03",
    "FAST_BOOK_DESCRIPTION_SOURCE",
    "FAST_BOOK_PART_II_III_2026_07_31",
    "FAST_BOOK_PART_II_III_SOURCE_URL",
    "PART_FUND_GROUPS",
    "TAS_COMPONENT_FORMAT_2026_08_03",
    "TAS_COMPONENT_FORMAT_SOURCE",
    "TREASURY_IDENTIFIER_AUTHORITY_URI",
    "TREASURY_PUBLISHER",
    "TREASURY_TAS_FAST_BOOK_GAPS",
    "AcquiredTreasuryPage",
    "AcquisitionMode",
    "FASTBookAccountRecord",
    "FASTBookPart",
    "FASTBookPublishedAccount",
    "FASTBookRecordError",
    "FASTBookWorkbookPin",
    "FetchedTreasuryPage",
    "FundGroup",
    "ParsedFASTBookDescription",
    "ParsedFASTBookWorkbook",
    "ParsedTASComponentFormat",
    "TASComponentError",
    "TASComponents",
    "TreasuryAcquisitionError",
    "TreasuryPageFetcher",
    "TreasuryPageSnapshotPin",
    "TreasuryPageSource",
    "TreasurySourceDriftError",
    "TreasuryTASFastBookEdition",
    "TreasuryTASFastBookError",
    "acquire_treasury_page",
    "assemble_treasury_tas_fast_book_edition",
    "fast_book_identifier",
    "parse_fast_book_description_page",
    "parse_fast_book_workbook",
    "parse_tas_canonical_value",
    "parse_tas_component_page",
    "parse_tas_components",
    "published_fast_book_identifier",
    "sha256_digest",
    "tas_identifier",
    "validate_fast_book_account_record",
]
