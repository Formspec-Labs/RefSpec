"""Pinned NAICS industry codes and PSC product/service codes.

The catalog decision for this source pair is explicit: NAICS and the Product
and Service Code (PSC) Manual are deterministic facets and optional ranking
signals. Neither states a document's policy topic, and no row here is ever
promoted to a general subject concept merely because it carries a readable
label.

Two independently governed publishers are pinned:

* The Census Bureau publishes NAICS as a five-year-cycle *vintage*. As of
  this capture the current published vintage is 2022; the next scheduled
  vintage is 2027, and Census had not published final 2027 structure data as
  of this capture (``naics_2027_available`` records that explicitly, it is
  never inferred). NAICS also uses hyphenated multi-sector ranges at the
  2-digit level (for example ``31-33`` for Manufacturing) that this module
  preserves verbatim rather than splitting or renumbering.
* Acquisition.gov publishes PSC as a numbered manual *edition* (for example
  "April 2025"). This module parses the publisher's Excel workbook and keeps
  the current four-character rows. Retired versions remain in the pinned
  source bytes and are excluded using their publisher-authored end dates.

Acquisition honesty for this pass: the official 2022 Census XLSX workbook is
captured and parsed in full. Acquisition.gov was unavailable during the PSC
capture. The Internet Archive's CDX index identified two April 2025 captures
of the exact official workbook URL with the same archive digest; the April 22
replay supplies the pinned publisher bytes. The original URL and replay URL
remain distinct provenance fields in the audit manifest.

Acquisition accepts a local exact capture or an injected fetcher. Importing
this module never opens a network connection.
"""

from __future__ import annotations

import csv
import hashlib
import io
import os
import re
import tempfile
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from types import MappingProxyType
from typing import Any, Literal, Protocol, cast
from urllib.parse import urlsplit

from openpyxl import load_workbook

from refspec.registry.infrastructure.controlled_identifier import ControlledIdentifier
from refspec.registry.infrastructure.source_controlled_resource import (
    ResourceUse,
    SourceControlledResourceBundle,
    build_source_controlled_resource_bundle,
)
from refspec.storage import canonical_json

NAICS_HOSTS = frozenset({"www.census.gov", "census.gov"})
PSC_HOSTS = frozenset({"www.acquisition.gov", "acquisition.gov"})
NAICS_IDENTIFIER_AUTHORITY_URI = "https://www.census.gov/naics/"
PSC_IDENTIFIER_AUTHORITY_URI = "https://www.acquisition.gov/psc-manual/all"
PSC_APRIL_2025_XLSX_URL = "https://www.acquisition.gov/sites/default/files/manual/PSC%20April%202025.xlsx"
PSC_APRIL_2025_WAYBACK_URL = (
    "https://web.archive.org/web/20250422004751id_/"
    "https://www.acquisition.gov/sites/default/files/manual/PSC%20April%202025.xlsx"
)
LANGUAGE = "en"

ResourceName = Literal["naicsCodes", "pscCodes"]
AcquisitionMode = Literal["cache", "local", "fetcher"]

_DIGEST = re.compile(r"^sha256:([0-9a-f]{64})$")
_NAICS_CODE = re.compile(r"^\d{2,6}$|^\d{2}-\d{2}$")
_PSC_CODE = re.compile(r"^[A-Z0-9]{4}$")
_PSC_CATEGORIES = frozenset({"Product", "Service", "Research and Development"})
_NAICS_FACET_BY_LENGTH = MappingProxyType(
    {
        2: "sector",
        3: "subsector",
        4: "industryGroup",
        5: "naicsIndustry",
        6: "nationalIndustry",
    }
)


class NaicsPscResourceError(ValueError):
    """Base class for NAICS/PSC controlled-code failures."""


class NaicsPscAcquisitionError(NaicsPscResourceError):
    """Exact official source bytes could not be acquired safely."""


class NaicsPscSourceDriftError(NaicsPscResourceError):
    """A NAICS or PSC source no longer matches the reviewed structure or pin."""


class NaicsPscAssignmentError(NaicsPscResourceError):
    """A record carries an unknown or inconsistent NAICS or PSC code."""


def sha256_digest(payload: bytes) -> str:
    """Return the canonical RefSpec SHA-256 spelling."""

    return "sha256:" + hashlib.sha256(payload).hexdigest()


@dataclass(frozen=True, slots=True)
class NaicsPscSource:
    """One official NAICS or PSC code-list rendering pinned by this module.

    ``edition`` carries the NAICS vintage year (for example "2022") or the
    PSC manual edition (for example "April 2025"); the field name is shared
    because both publishers version their code list the same way -- one
    dated release supersedes the previous one in place.
    """

    resource_name: ResourceName
    title: str
    source_url: str
    hosts: frozenset[str]
    filename: str
    expected_count: int
    edition: str

    def __post_init__(self) -> None:
        parsed = urlsplit(self.source_url)
        if parsed.scheme != "https" or parsed.hostname not in self.hosts:
            raise NaicsPscAcquisitionError("source_url must be an official HTTPS URL on the declared host set")
        if parsed.username is not None or parsed.password is not None:
            raise NaicsPscAcquisitionError("source_url must not contain credentials")
        if not self.filename or Path(self.filename).name != self.filename:
            raise NaicsPscAcquisitionError("filename must be one plain path component")
        if self.expected_count <= 0:
            raise NaicsPscAcquisitionError("expected_count must be positive")
        if not self.edition.strip():
            raise NaicsPscAcquisitionError("edition must not be empty")


NAICS_CODES_SOURCE = NaicsPscSource(
    resource_name="naicsCodes",
    title="2022 NAICS US Structure",
    source_url="https://www.census.gov/naics/",
    hosts=NAICS_HOSTS,
    filename="naics-2022-us-structure.csv",
    expected_count=14,
    edition="2022",
)
NAICS_CODES_XLSX_SOURCE = NaicsPscSource(
    resource_name="naicsCodes",
    title="2022 NAICS US Structure",
    source_url="https://www.census.gov/naics/2022NAICS/2-6%20digit_2022_Codes.xlsx",
    hosts=NAICS_HOSTS,
    filename="2-6-digit_2022_Codes.xlsx",
    expected_count=2_125,
    edition="2022",
)
PSC_CODES_SOURCE = NaicsPscSource(
    resource_name="pscCodes",
    title="Product and Service Code Manual",
    source_url="https://www.acquisition.gov/psc-manual/all",
    hosts=PSC_HOSTS,
    filename="psc-manual-april-2025.csv",
    expected_count=8,
    edition="April 2025",
)
PSC_CODES_XLSX_SOURCE = NaicsPscSource(
    resource_name="pscCodes",
    title="Product and Service Code Manual",
    source_url=PSC_APRIL_2025_XLSX_URL,
    hosts=PSC_HOSTS,
    filename="PSC-April-2025.xlsx",
    expected_count=2_344,
    edition="April 2025",
)

# NAICS 2027 is the next scheduled five-year revision cycle. Census had not
# published final 2027 structure data as of this capture; this constant
# records that gap explicitly rather than a caller inferring it from an
# absent resource.
NAICS_2027_STRUCTURE_PUBLISHED = False


@dataclass(frozen=True, slots=True)
class NaicsPscSnapshotPin:
    """Exact identity of one captured NAICS or PSC code-list rendering."""

    source: NaicsPscSource
    retrieved_at: str
    expected_sha256: str
    expected_byte_length: int

    def __post_init__(self) -> None:
        if _DIGEST.fullmatch(self.expected_sha256) is None:
            raise NaicsPscAcquisitionError("expected_sha256 must be a lowercase sha256:<64 hex> digest")
        if self.expected_byte_length <= 0:
            raise NaicsPscAcquisitionError("expected_byte_length must be positive")
        if not self.retrieved_at.strip():
            raise NaicsPscAcquisitionError("retrieved_at must not be empty")


# These two pins are constructed CSV fixtures retained for narrow parser tests.
NAICS_CODES_2026_08_03 = NaicsPscSnapshotPin(
    source=NAICS_CODES_SOURCE,
    retrieved_at="2026-08-03T20:00:00Z",
    expected_sha256="sha256:a8e7fb37571ba8c2e7eb8281e2805f9f4a3ef77104fed3c3bd43e7be072c3539",
    expected_byte_length=517,
)
PSC_CODES_2026_08_03 = NaicsPscSnapshotPin(
    source=PSC_CODES_SOURCE,
    retrieved_at="2026-08-03T20:00:00Z",
    expected_sha256="sha256:dd6c5307bb761b842152ed91d20be99943b890c38c4c1f92ae15cb60b3dc9ba5",
    expected_byte_length=545,
)
NAICS_CODES_2022_XLSX = NaicsPscSnapshotPin(
    source=NAICS_CODES_XLSX_SOURCE,
    retrieved_at="2026-08-03T20:00:00Z",
    expected_sha256="sha256:be12ba41002803359f49181c9bf33a03fbd08578f4f4a4c0bbad7aadaaea0316",
    expected_byte_length=82_460,
)
PSC_CODES_APRIL_2025_XLSX = NaicsPscSnapshotPin(
    source=PSC_CODES_XLSX_SOURCE,
    retrieved_at="2026-08-04T01:17:18Z",
    expected_sha256="sha256:5ae8159d8dff645f24e5b397decc4914f7efebb25f7777cbea8e75ab7e8430f4",
    expected_byte_length=462_762,
)


@dataclass(frozen=True, slots=True)
class FetchedNaicsPscResponse:
    """Provider-independent response returned by an injected fetcher."""

    body: bytes
    status_code: int
    content_type: str
    resolved_url: str


class NaicsPscFetcher(Protocol):
    """Small transport boundary for the official NAICS and PSC sources."""

    def fetch(self, source_url: str, *, timeout_seconds: float) -> FetchedNaicsPscResponse:
        """Fetch one response while preserving its exact body bytes."""


@dataclass(frozen=True, slots=True)
class AcquiredNaicsPscSource:
    """One verified source object in the content-addressed store."""

    pin: NaicsPscSnapshotPin
    path: Path
    sha256: str
    byte_length: int
    source_url: str
    resolved_url: str | None
    content_type: str
    acquisition_mode: AcquisitionMode
    cache_hit: bool
    local_source_path: Path | None


def _validate_resolved_url(value: str, hosts: frozenset[str]) -> None:
    parsed = urlsplit(value)
    if parsed.scheme != "https" or parsed.hostname not in hosts:
        raise NaicsPscAcquisitionError("fetcher resolved_url must remain on the official HTTPS source host")
    if parsed.username is not None or parsed.password is not None:
        raise NaicsPscAcquisitionError("fetcher resolved_url must not contain credentials")


def _verify_payload(payload: bytes, pin: NaicsPscSnapshotPin, *, location: str) -> tuple[str, int]:
    byte_length = len(payload)
    if byte_length != pin.expected_byte_length:
        raise NaicsPscSourceDriftError(
            f"{location} byte length drift: expected {pin.expected_byte_length}, got {byte_length}"
        )
    actual_sha256 = sha256_digest(payload)
    if actual_sha256 != pin.expected_sha256:
        raise NaicsPscSourceDriftError(f"{location} digest drift: expected {pin.expected_sha256}, got {actual_sha256}")
    if pin.source.filename.endswith(".csv"):
        try:
            payload.decode("utf-8")
        except UnicodeDecodeError as error:
            raise NaicsPscSourceDriftError(f"{location} is not valid UTF-8 text") from error
    elif not pin.source.filename.endswith(".xlsx") or not payload.startswith(b"PK"):
        raise NaicsPscSourceDriftError(f"{location} is not the declared CSV or XLSX source")
    return actual_sha256, byte_length


def _verify_existing(path: Path, pin: NaicsPscSnapshotPin) -> AcquiredNaicsPscSource:
    if path.is_symlink() or not path.is_file():
        raise NaicsPscAcquisitionError(f"content-addressed target is not a regular file: {path}")
    actual_sha256, byte_length = _verify_payload(
        path.read_bytes(),
        pin,
        location="cached NAICS/PSC source",
    )
    return AcquiredNaicsPscSource(
        pin=pin,
        path=path,
        sha256=actual_sha256,
        byte_length=byte_length,
        source_url=pin.source.source_url,
        resolved_url=None,
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            if pin.source.filename.endswith(".xlsx")
            else "text/csv"
        ),
        acquisition_mode="cache",
        cache_hit=True,
        local_source_path=None,
    )


def _publish_payload(
    payload: bytes,
    pin: NaicsPscSnapshotPin,
    final_path: Path,
    *,
    content_type: str,
    acquisition_mode: Literal["local", "fetcher"],
    resolved_url: str | None,
    local_source_path: Path | None,
) -> AcquiredNaicsPscSource:
    actual_sha256, byte_length = _verify_payload(
        payload,
        pin,
        location=f"{acquisition_mode} NAICS/PSC source",
    )
    final_path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=".acquire-",
        suffix=".tmp",
        dir=final_path.parent,
    )
    temporary_path = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as output:
            descriptor = -1
            output.write(payload)
            output.flush()
            os.fsync(output.fileno())
        try:
            os.link(temporary_path, final_path)
        except FileExistsError:
            return _verify_existing(final_path, pin)
        return AcquiredNaicsPscSource(
            pin=pin,
            path=final_path,
            sha256=actual_sha256,
            byte_length=byte_length,
            source_url=pin.source.source_url,
            resolved_url=resolved_url,
            content_type=content_type,
            acquisition_mode=acquisition_mode,
            cache_hit=False,
            local_source_path=local_source_path,
        )
    finally:
        if descriptor >= 0:
            os.close(descriptor)
        temporary_path.unlink(missing_ok=True)


def acquire_naics_psc_source(
    pin: NaicsPscSnapshotPin,
    store_dir: Path,
    *,
    source_path: Path | None = None,
    fetcher: NaicsPscFetcher | None = None,
    timeout_seconds: float = 30.0,
) -> AcquiredNaicsPscSource:
    """Acquire one exact NAICS or PSC rendering through a provider-neutral boundary."""

    if timeout_seconds <= 0:
        raise NaicsPscAcquisitionError("timeout_seconds must be positive")
    if source_path is not None and fetcher is not None:
        raise NaicsPscAcquisitionError("provide source_path or fetcher, not both")
    digest_hex = cast(re.Match[str], _DIGEST.fullmatch(pin.expected_sha256)).group(1)
    final_path = Path(store_dir) / "sha256" / digest_hex / pin.source.filename
    if final_path.exists() or final_path.is_symlink():
        return _verify_existing(final_path, pin)

    if source_path is not None:
        local_path = Path(source_path)
        if local_path.is_symlink() or not local_path.is_file():
            raise NaicsPscAcquisitionError(f"local NAICS/PSC source is not a regular file: {local_path}")
        return _publish_payload(
            local_path.read_bytes(),
            pin,
            final_path,
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                if pin.source.filename.endswith(".xlsx")
                else "text/csv"
            ),
            acquisition_mode="local",
            resolved_url=None,
            local_source_path=local_path.resolve(),
        )

    if fetcher is None:
        raise NaicsPscAcquisitionError("NAICS/PSC source is not cached; provide source_path or an injected fetcher")
    fetched = fetcher.fetch(pin.source.source_url, timeout_seconds=timeout_seconds)
    if fetched.status_code != 200:
        raise NaicsPscAcquisitionError(f"could not acquire {pin.source.source_url}: HTTP {fetched.status_code}")
    _validate_resolved_url(fetched.resolved_url, pin.source.hosts)
    media_type = fetched.content_type.partition(";")[0].strip().lower()
    if media_type not in {
        "text/csv",
        "application/csv",
        "text/plain",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }:
        raise NaicsPscSourceDriftError(f"{pin.source.resource_name} content type drifted to {fetched.content_type!r}")
    return _publish_payload(
        fetched.body,
        pin,
        final_path,
        content_type=fetched.content_type,
        acquisition_mode="fetcher",
        resolved_url=fetched.resolved_url,
        local_source_path=None,
    )


@dataclass(frozen=True, slots=True)
class NaicsPscCode:
    """One exact publisher code and label, plus the facet computed for it.

    ``facet`` carries the NAICS hierarchy level (``sector``, ``subsector``,
    ``industryGroup``, ``naicsIndustry``, or ``nationalIndustry``), derived
    from the code's own digit shape, or the PSC category (``Product``,
    ``Service``, or ``Research and Development``) read verbatim from the
    source's own Category column. Neither facet is ever treated as a
    document's policy topic.
    """

    resource_name: ResourceName
    use: ResourceUse
    publisher_label: str
    source_url: str
    identifiers: tuple[ControlledIdentifier, ...]
    facet: str
    is_general_subject_concept: bool = False


@dataclass(frozen=True, slots=True)
class ParsedNaicsPscResource:
    """A parsed, digest-pinned NAICS or PSC code list."""

    source: NaicsPscSource
    retrieved_at: str
    source_sha256: str
    source_byte_length: int
    edition: str
    codes: tuple[NaicsPscCode, ...]
    gaps: tuple[str, ...]

    def by_code(self) -> dict[str, NaicsPscCode]:
        """Index the endpoint's publisher-issued code, retaining every other ID."""

        code_kind = "naicsCode" if self.source.resource_name == "naicsCodes" else "pscCode"
        result: dict[str, NaicsPscCode] = {}
        for entry in self.codes:
            matches = [identifier for identifier in entry.identifiers if identifier.kind == code_kind]
            if len(matches) != 1:
                raise NaicsPscSourceDriftError(f"{self.source.resource_name} row must retain exactly one {code_kind}")
            result[matches[0].value] = entry
        return result


def _read_acquired_payload(acquired: AcquiredNaicsPscSource) -> bytes:
    payload = acquired.path.read_bytes()
    _verify_payload(payload, acquired.pin, location="parsed NAICS/PSC source")
    return payload


def _csv_rows(payload: bytes) -> list[list[str]]:
    decoded = payload.decode("utf-8")
    return list(csv.reader(io.StringIO(decoded)))


def _naics_xlsx_rows(payload: bytes) -> list[list[str]]:
    """Read the publisher workbook's first three populated columns."""

    try:
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    except (OSError, ValueError) as error:
        raise NaicsPscSourceDriftError("NAICS XLSX workbook is unreadable") from error
    try:
        if len(workbook.sheetnames) != 1:
            raise NaicsPscSourceDriftError("NAICS XLSX workbook must contain exactly one worksheet")
        rows: list[list[str]] = []
        for raw_row in workbook.active.iter_rows(values_only=True):
            values = raw_row[:3]
            if all(value is None for value in values):
                continue
            if any(value is None for value in values):
                raise NaicsPscSourceDriftError("NAICS XLSX populated row has a blank required cell")
            rows.append([str(value).strip() for value in values])
        return rows
    finally:
        workbook.close()


_PSC_XLSX_HEADER = (
    "PSC CODE",
    "PRODUCT AND SERVICE CODE NAME",
    "START DATE",
    "END DATE",
    "PRODUCT AND SERVICE CODE FULL NAME (DESCRIPTION)",
    "PRODUCT AND SERVICE CODE INCLUDES",
    "PRODUCT AND SERVICE CODE EXCLUDES",
    "PRODUCT AND SERVICE CODE NOTES",
    "Parent PSC Code",
    "PSC Category: Service (S)/Product (P)",
    "Level 1 Category Code",
    "Level 1 Category",
    "Level 2 Category Code",
    "Level 2 Category",
)
_PSC_XLSX_SOURCE_ROW_COUNT = 6_108


def _psc_code_cell(value: object) -> str:
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip() if value is not None else ""


def _psc_date(value: object, label: str) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raise NaicsPscSourceDriftError(f"{label} must be an Excel date")


def _parse_psc_xlsx(
    acquired: AcquiredNaicsPscSource,
    payload: bytes,
) -> ParsedNaicsPscResource:
    source = acquired.pin.source
    try:
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    except (OSError, ValueError) as error:
        raise NaicsPscSourceDriftError("PSC XLSX workbook is unreadable") from error
    try:
        if workbook.sheetnames != ["PSC for 042025", "Category Managers"]:
            raise NaicsPscSourceDriftError(f"PSC XLSX worksheets drifted: {workbook.sheetnames}")
        worksheet = workbook["PSC for 042025"]
        worksheet.calculate_dimension(force=True)
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows or tuple(rows[0][:14]) != _PSC_XLSX_HEADER:
            raise NaicsPscSourceDriftError("PSC XLSX header drifted")
        data_rows = rows[1:]
        if len(data_rows) != _PSC_XLSX_SOURCE_ROW_COUNT:
            raise NaicsPscSourceDriftError(
                f"PSC XLSX source row count drift: expected {_PSC_XLSX_SOURCE_ROW_COUNT}, got {len(data_rows)}"
            )

        codes: list[NaicsPscCode] = []
        seen_codes: set[str] = set()
        for source_row, row in enumerate(data_rows, start=2):
            code = _psc_code_cell(row[0])
            if _PSC_CODE.fullmatch(code) is None or row[3] is not None:
                continue
            label = row[1]
            if not isinstance(label, str) or not label.strip() or label != label.strip():
                raise NaicsPscSourceDriftError(f"PSC XLSX row {source_row} has a malformed publisher label")
            if code in seen_codes:
                raise NaicsPscSourceDriftError(f"active PSC code {code!r} is duplicated")
            seen_codes.add(code)
            level_one_category = row[11]
            parent_category = row[8]
            raw_facet = level_one_category if level_one_category is not None else parent_category
            if not isinstance(raw_facet, str) or not raw_facet.strip():
                raise NaicsPscSourceDriftError(f"PSC XLSX row {source_row} has no publisher category")
            codes.append(
                NaicsPscCode(
                    resource_name="pscCodes",
                    use="deterministicMetadata",
                    publisher_label=label,
                    source_url=source.source_url,
                    identifiers=(
                        ControlledIdentifier(
                            value=code,
                            kind="pscCode",
                            authority_uri=PSC_IDENTIFIER_AUTHORITY_URI,
                            source_uri=source.source_url,
                            observed_at=acquired.pin.retrieved_at,
                            effective_at=_psc_date(row[2], f"PSC XLSX row {source_row} START DATE"),
                            source_digest=acquired.sha256,
                        ),
                    ),
                    facet=raw_facet.strip(),
                )
            )
        if len(codes) != source.expected_count:
            raise NaicsPscSourceDriftError(
                f"pscCodes count drift: expected {source.expected_count}, parsed {len(codes)}"
            )
        return ParsedNaicsPscResource(
            source=source,
            retrieved_at=acquired.pin.retrieved_at,
            source_sha256=acquired.sha256,
            source_byte_length=acquired.byte_length,
            edition=source.edition,
            codes=tuple(codes),
            gaps=NAICS_PSC_PORTFOLIO_GAPS,
        )
    finally:
        workbook.close()


def _naics_facet(code: str) -> str:
    if "-" in code:
        first, _, second = code.partition("-")
        if len(first) != 2 or len(second) != 2 or not first.isdigit() or not second.isdigit() or first >= second:
            raise NaicsPscSourceDriftError(f"NAICS sector range {code!r} is not an ascending two-digit range")
        return "sector"
    facet = _NAICS_FACET_BY_LENGTH.get(len(code))
    if facet is None:
        raise NaicsPscSourceDriftError(f"NAICS code {code!r} has an unsupported digit length")
    return facet


def parse_naics_codes(acquired: AcquiredNaicsPscSource) -> ParsedNaicsPscResource:
    """Parse the exact NAICS US Structure rows without minting subjects."""

    source = acquired.pin.source
    if source.resource_name != "naicsCodes":
        raise NaicsPscSourceDriftError("acquired source was not pinned against a NAICS codes source")
    payload = _read_acquired_payload(acquired)
    rows = _naics_xlsx_rows(payload) if source.filename.endswith(".xlsx") else _csv_rows(payload)
    if not rows:
        raise NaicsPscSourceDriftError("NAICS payload has no rows")
    expected_header = ["Seq. No.", f"{source.edition} NAICS US Code", f"{source.edition} NAICS US Title"]
    normalized_header = [" ".join(value.split()) for value in rows[0]]
    if normalized_header != expected_header:
        raise NaicsPscSourceDriftError(f"NAICS header drifted: expected {expected_header}, got {rows[0]}")
    data_rows = rows[1:]
    if len(data_rows) != source.expected_count:
        raise NaicsPscSourceDriftError(
            f"naicsCodes count drift: expected {source.expected_count}, parsed {len(data_rows)}"
        )

    codes: list[NaicsPscCode] = []
    seen_codes: set[str] = set()
    for ordinal, row in enumerate(data_rows, start=1):
        if len(row) != 3:
            raise NaicsPscSourceDriftError(f"NAICS row {ordinal} must contain exactly three fields")
        seq_text, code, label = row
        if seq_text.strip() != str(ordinal):
            raise NaicsPscSourceDriftError(f"NAICS row {ordinal} has out-of-sequence Seq. No. {seq_text!r}")
        if _NAICS_CODE.fullmatch(code) is None:
            raise NaicsPscSourceDriftError(f"NAICS row {ordinal} has a malformed code {code!r}")
        if not label or label != label.strip():
            raise NaicsPscSourceDriftError(f"NAICS row {ordinal} has a malformed publisher label")
        if code in seen_codes:
            raise NaicsPscSourceDriftError(f"NAICS code {code!r} is duplicated")
        seen_codes.add(code)
        codes.append(
            NaicsPscCode(
                resource_name="naicsCodes",
                use="deterministicMetadata",
                publisher_label=label,
                source_url=source.source_url,
                identifiers=(
                    ControlledIdentifier(
                        value=code,
                        kind="naicsCode",
                        authority_uri=NAICS_IDENTIFIER_AUTHORITY_URI,
                        source_uri=source.source_url,
                        observed_at=acquired.pin.retrieved_at,
                        effective_at=None,
                        source_digest=acquired.sha256,
                    ),
                ),
                facet=_naics_facet(code),
            )
        )

    return ParsedNaicsPscResource(
        source=source,
        retrieved_at=acquired.pin.retrieved_at,
        source_sha256=acquired.sha256,
        source_byte_length=acquired.byte_length,
        edition=source.edition,
        codes=tuple(codes),
        gaps=NAICS_PSC_PORTFOLIO_GAPS,
    )


def parse_psc_codes(acquired: AcquiredNaicsPscSource) -> ParsedNaicsPscResource:
    """Parse the exact PSC Manual rows without minting subjects."""

    source = acquired.pin.source
    if source.resource_name != "pscCodes":
        raise NaicsPscSourceDriftError("acquired source was not pinned against a PSC codes source")
    payload = _read_acquired_payload(acquired)
    if source.filename.endswith(".xlsx"):
        return _parse_psc_xlsx(acquired, payload)
    rows = _csv_rows(payload)
    if not rows:
        raise NaicsPscSourceDriftError("PSC payload has no rows")
    expected_header = ["PSC Code", "PSC Name", "Category", "Manual Edition"]
    if rows[0] != expected_header:
        raise NaicsPscSourceDriftError(f"PSC header drifted: expected {expected_header}, got {rows[0]}")
    data_rows = rows[1:]
    if len(data_rows) != source.expected_count:
        raise NaicsPscSourceDriftError(
            f"pscCodes count drift: expected {source.expected_count}, parsed {len(data_rows)}"
        )

    codes: list[NaicsPscCode] = []
    seen_codes: set[str] = set()
    for ordinal, row in enumerate(data_rows, start=1):
        if len(row) != 4:
            raise NaicsPscSourceDriftError(f"PSC row {ordinal} must contain exactly four fields")
        code, label, category, edition = row
        if _PSC_CODE.fullmatch(code) is None:
            raise NaicsPscSourceDriftError(f"PSC row {ordinal} has a malformed code {code!r}")
        if not label or label != label.strip():
            raise NaicsPscSourceDriftError(f"PSC row {ordinal} has a malformed publisher label")
        if category not in _PSC_CATEGORIES:
            raise NaicsPscSourceDriftError(f"PSC row {ordinal} has an unknown Category {category!r}")
        if edition != source.edition:
            raise NaicsPscSourceDriftError(
                f"PSC row {ordinal} Manual Edition {edition!r} does not match pinned edition {source.edition!r}"
            )
        if code in seen_codes:
            raise NaicsPscSourceDriftError(f"PSC code {code!r} is duplicated")
        seen_codes.add(code)
        codes.append(
            NaicsPscCode(
                resource_name="pscCodes",
                use="deterministicMetadata",
                publisher_label=label,
                source_url=source.source_url,
                identifiers=(
                    ControlledIdentifier(
                        value=code,
                        kind="pscCode",
                        authority_uri=PSC_IDENTIFIER_AUTHORITY_URI,
                        source_uri=source.source_url,
                        observed_at=acquired.pin.retrieved_at,
                        effective_at=None,
                        source_digest=acquired.sha256,
                    ),
                ),
                facet=category,
            )
        )

    return ParsedNaicsPscResource(
        source=source,
        retrieved_at=acquired.pin.retrieved_at,
        source_sha256=acquired.sha256,
        source_byte_length=acquired.byte_length,
        edition=source.edition,
        codes=tuple(codes),
        gaps=NAICS_PSC_PORTFOLIO_GAPS,
    )


NAICS_PSC_PORTFOLIO_GAPS = (
    (
        "NAICS and PSC codes are deterministic facets and optional ranking "
        "signals; they do not state a document's policy topic and are never "
        "promoted to a general subject concept."
    ),
    (
        "The Census Bureau revises NAICS on a five-year cycle (2017, 2022, "
        "2027); this module packages only the 2022 US structure vintage. "
        "The 2027 vintage had not been published as final Census structure "
        "data as of this capture -- see NAICS_2027_STRUCTURE_PUBLISHED."
    ),
    (
        "The PSC Manual April 2025 workbook contains active and retired code "
        "versions. This module imports the 2,344 current four-character rows "
        "whose publisher-authored END DATE is blank; all 6,108 source rows "
        "remain preserved in the pinned workbook bytes."
    ),
    (
        "Acquisition.gov was unavailable during capture. The PSC workbook "
        "was recovered from the Internet Archive's April 22, 2025 replay of "
        "the exact official acquisition.gov XLSX URL; archive metadata and "
        "the original URL are retained separately in the source manifest."
    ),
)


@dataclass(frozen=True, slots=True)
class NaicsPscPortfolio:
    """The two imported resources and their documented capture gaps."""

    naics_codes: ParsedNaicsPscResource
    psc_codes: ParsedNaicsPscResource
    gaps: tuple[str, ...]


def assemble_naics_psc_portfolio(
    resources: Sequence[ParsedNaicsPscResource],
) -> NaicsPscPortfolio:
    """Require both distinct resources and retain the documented gaps."""

    by_name = {resource.source.resource_name: resource for resource in resources}
    if len(resources) != 2 or set(by_name) != {"naicsCodes", "pscCodes"}:
        raise NaicsPscSourceDriftError("NAICS/PSC portfolio requires exactly one NAICS and one PSC resource")
    return NaicsPscPortfolio(
        naics_codes=by_name["naicsCodes"],
        psc_codes=by_name["pscCodes"],
        gaps=NAICS_PSC_PORTFOLIO_GAPS,
    )


@dataclass(frozen=True, slots=True)
class NaicsPscAssignment:
    """One field's value validated against its exact pinned code list."""

    field: str
    publisher_label: str
    use: ResourceUse
    identifiers: tuple[ControlledIdentifier, ...]
    facet: str
    is_general_subject_concept: bool


@dataclass(frozen=True, slots=True)
class ValidatedNaicsPscClassification:
    """Optional NAICS/PSC facets validated against one procurement record.

    ``record_reference`` is preserved verbatim and is never replaced; the
    NAICS and PSC assignments are optional, independently validated
    secondary facets, matching the catalog's deterministic-facet decision.
    """

    record_reference: str
    naics: NaicsPscAssignment | None
    psc: NaicsPscAssignment | None
    gaps: tuple[str, ...]


def _assignment(field: str, code: NaicsPscCode) -> NaicsPscAssignment:
    return NaicsPscAssignment(
        field=field,
        publisher_label=code.publisher_label,
        use=code.use,
        identifiers=code.identifiers,
        facet=code.facet,
        is_general_subject_concept=code.is_general_subject_concept,
    )


def validate_naics_psc_classification(
    record: Mapping[str, object],
    portfolio: NaicsPscPortfolio,
) -> ValidatedNaicsPscClassification:
    """Validate an optional procurement record's NAICS/PSC facets, failing closed.

    ``record_reference`` must remain the record's own native identity;
    ``naics_code`` and ``psc_code`` are optional and, when present, must
    match a pinned code exactly or this refuses the record.
    """

    raw_reference = record.get("record_reference")
    if not isinstance(raw_reference, str) or not raw_reference.strip():
        raise NaicsPscAssignmentError("record must carry a non-empty record_reference")

    naics_assignment: NaicsPscAssignment | None = None
    raw_naics = record.get("naics_code")
    if raw_naics is not None:
        if not isinstance(raw_naics, str):
            raise NaicsPscAssignmentError("naics_code must be a string")
        code = portfolio.naics_codes.by_code().get(raw_naics)
        if code is None:
            raise NaicsPscAssignmentError(f"unknown NAICS code {raw_naics!r}")
        naics_assignment = _assignment("naics_code", code)

    psc_assignment: NaicsPscAssignment | None = None
    raw_psc = record.get("psc_code")
    if raw_psc is not None:
        if not isinstance(raw_psc, str):
            raise NaicsPscAssignmentError("psc_code must be a string")
        code = portfolio.psc_codes.by_code().get(raw_psc)
        if code is None:
            raise NaicsPscAssignmentError(f"unknown PSC code {raw_psc!r}")
        psc_assignment = _assignment("psc_code", code)

    return ValidatedNaicsPscClassification(
        record_reference=raw_reference,
        naics=naics_assignment,
        psc=psc_assignment,
        gaps=NAICS_PSC_PORTFOLIO_GAPS,
    )


_DETERMINISTIC_FACET_ROLE_GAP = MappingProxyType(
    {
        "kind": "deterministicFacetRole",
        "reason": (
            "NAICS and PSC codes are deterministic facets and optional "
            "ranking signals; they do not state a document's policy topic "
            "and are never promoted to a general subject concept."
        ),
    }
)
_NAICS_VINTAGE_GAP = MappingProxyType(
    {
        "kind": "naicsVintageUnavailable",
        "reason": (
            "The 2027 NAICS vintage had not been published as final Census "
            "structure data as of this capture; only the 2022 US structure "
            "vintage is packaged."
        ),
    }
)
_PSC_BINARY_MANUAL_GAP = MappingProxyType(
    {
        "kind": "pscManualBinaryFormat",
        "reason": (
            "The constructed PSC CSV fixture covers only a small parser "
            "sample; use the pinned April 2025 XLSX source for real-data use."
        ),
    }
)
_UNVERIFIED_LIVE_CAPTURE_GAP = MappingProxyType(
    {
        "kind": "unverifiedLiveCapture",
        "reason": (
            "Neither publisher's code-list bytes were captured live this "
            "session; the packaged source artifact is a constructed "
            "fixture, not a verified live capture."
        ),
    }
)


def _verified_source_payload(acquired: AcquiredNaicsPscSource) -> bytes:
    payload = acquired.path.read_bytes()
    if len(payload) != acquired.byte_length or sha256_digest(payload) != acquired.sha256:
        raise NaicsPscSourceDriftError("NAICS/PSC package source differs from its acquired pin")
    return payload


def _observation(
    resource_id: str,
    parsed: ParsedNaicsPscResource,
    ordinal: int,
    code: NaicsPscCode,
) -> dict[str, Any]:
    identifiers = [
        {
            "value": identifier.value,
            "kind": identifier.kind,
            "authorityUri": identifier.authority_uri,
            "sourceUri": identifier.source_uri,
            "sourcePath": f"$.rows[{ordinal}]",
            "observedAt": identifier.observed_at,
            "sourceDigest": identifier.source_digest,
        }
        for identifier in code.identifiers
    ]
    identity = {
        "resourceId": resource_id,
        "sourceArtifact": parsed.source.source_url,
        "sourceOrdinal": ordinal,
        "publisherLabel": code.publisher_label,
        "identifiers": identifiers,
    }
    digest = hashlib.sha256(canonical_json(identity).encode("utf-8")).hexdigest()
    return {
        "id": f"urn:ref:source-observation:{resource_id}:{digest}",
        "sourceArtifact": parsed.source.source_url,
        "sourcePath": f"$.rows[{ordinal}]",
        "sourceOrdinal": ordinal,
        "labels": [{"value": code.publisher_label, "language": LANGUAGE, "role": "preferred"}],
        "identifiers": identifiers,
        "uses": ["deterministicMetadata"],
        "conceptIdentityClaimed": False,
    }


def _build_package(
    *,
    resource_id: str,
    title: str,
    acquired: AcquiredNaicsPscSource,
    parsed: ParsedNaicsPscResource,
    package_gaps: Sequence[Mapping[str, Any]],
) -> SourceControlledResourceBundle:
    if parsed.source_sha256 != acquired.sha256 or parsed.source.source_url != acquired.pin.source.source_url:
        raise NaicsPscSourceDriftError("parsed resource and acquired source describe different sources")
    payload = _verified_source_payload(acquired)
    observations = tuple(_observation(resource_id, parsed, ordinal, code) for ordinal, code in enumerate(parsed.codes))
    return build_source_controlled_resource_bundle(
        resource_id=resource_id,
        title=title,
        resource_kind="controlledCodeList",
        identity_status="publisherIdentifiersPreserved",
        uses=("deterministicMetadata",),
        captured_at=parsed.retrieved_at,
        observations=observations,
        source_artifacts={parsed.source.source_url: payload},
        source_observed_count=parsed.source.expected_count,
        gaps=(
            _DETERMINISTIC_FACET_ROLE_GAP,
            *((_UNVERIFIED_LIVE_CAPTURE_GAP,) if acquired.pin.source.filename.endswith(".csv") else ()),
            *package_gaps,
        ),
    )


def build_naics_code_package(
    acquired: AcquiredNaicsPscSource,
    parsed: ParsedNaicsPscResource,
) -> SourceControlledResourceBundle:
    """Package all exact NAICS 2022 US Structure codes as deterministic facets."""

    return _build_package(
        resource_id="naics-2022-us-structure-2026-08-03",
        title="NAICS 2022 US Structure, vintage 2022",
        acquired=acquired,
        parsed=parsed,
        package_gaps=(_NAICS_VINTAGE_GAP,),
    )


def build_psc_code_package(
    acquired: AcquiredNaicsPscSource,
    parsed: ParsedNaicsPscResource,
) -> SourceControlledResourceBundle:
    """Package all exact PSC Manual codes as deterministic facets."""

    return _build_package(
        resource_id="psc-manual-april-2025-2026-08-03",
        title="Product and Service Code Manual, edition April 2025",
        acquired=acquired,
        parsed=parsed,
        package_gaps=((_PSC_BINARY_MANUAL_GAP,) if acquired.pin.source.filename.endswith(".csv") else ()),
    )


__all__ = [
    "NAICS_2027_STRUCTURE_PUBLISHED",
    "NAICS_CODES_2022_XLSX",
    "NAICS_CODES_2026_08_03",
    "NAICS_CODES_SOURCE",
    "NAICS_CODES_XLSX_SOURCE",
    "NAICS_HOSTS",
    "NAICS_IDENTIFIER_AUTHORITY_URI",
    "NAICS_PSC_PORTFOLIO_GAPS",
    "PSC_APRIL_2025_WAYBACK_URL",
    "PSC_APRIL_2025_XLSX_URL",
    "PSC_CODES_2026_08_03",
    "PSC_CODES_APRIL_2025_XLSX",
    "PSC_CODES_SOURCE",
    "PSC_CODES_XLSX_SOURCE",
    "PSC_HOSTS",
    "PSC_IDENTIFIER_AUTHORITY_URI",
    "AcquiredNaicsPscSource",
    "AcquisitionMode",
    "FetchedNaicsPscResponse",
    "NaicsPscAcquisitionError",
    "NaicsPscAssignment",
    "NaicsPscAssignmentError",
    "NaicsPscCode",
    "NaicsPscFetcher",
    "NaicsPscPortfolio",
    "NaicsPscResourceError",
    "NaicsPscSnapshotPin",
    "NaicsPscSource",
    "NaicsPscSourceDriftError",
    "ParsedNaicsPscResource",
    "ValidatedNaicsPscClassification",
    "acquire_naics_psc_source",
    "assemble_naics_psc_portfolio",
    "build_naics_code_package",
    "build_psc_code_package",
    "parse_naics_codes",
    "parse_psc_codes",
    "sha256_digest",
    "validate_naics_psc_classification",
]
