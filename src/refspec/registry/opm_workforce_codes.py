"""Pinned OPM workforce and PLUM position/status code imports.

OPM's Federal Workforce Data downloads
(https://data.opm.gov/explore-data/data/data-downloads) and the PLUM Act
position data release
(https://www.opm.gov/about-us/open-government/plum-reporting/plum-data/)
publish occupational series, pay plan, appointment type, work schedule, and
PLUM appointment-authority/incumbent-status codes. These are entity and
observation metadata code sets, not a document-topic vocabulary: an
occupational series or pay plan code never states what a document is about.

The current real-data path parses OPM's complete three-sheet EHRI workbook
export and PLUM all-data CSV. The older five-resource package definitions below
remain small, documented-shape development samples for compatibility; they are
not evidence for the real-data gate. Their strict counts and digests prevent a
real export from being silently mistaken for one of those legacy samples.

PLUM position data additionally carries statutory redaction, agency
certification, and release-vintage rules: some incumbent identities may be
withheld for protected positions, each release must be certified by the
submitting agency, and a record must be checked against the exact edition
(vintage) whose codes were pinned. This module encodes those rules as
refusal-style checks on downstream observation records; it never ingests
bulk PLUM position rows, only the small, closed appointment-authority and
incumbent-status-marker vocabulary.

Acquisition accepts a local exact capture or an injected fetcher. Importing
this module never opens a network connection.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import re
import tempfile
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path
from types import MappingProxyType
from typing import Any, Literal, Protocol, cast
from urllib.parse import urlsplit

from openpyxl import load_workbook

from refspec.registry.infrastructure.controlled_identifier import ControlledIdentifier
from refspec.registry.infrastructure.pinned_acquisition import FetcherAcquisitionMode as AcquisitionMode
from refspec.registry.infrastructure.source_controlled_resource import (
    ResourceUse as PackageResourceUse,
)
from refspec.registry.infrastructure.source_controlled_resource import (
    SourceControlledResourceBundle,
    SourceControlledResourceView,
    build_source_controlled_resource_bundle,
)
from refspec.storage import canonical_json

OPM_PUBLISHER = "U.S. Office of Personnel Management"
OPM_IDENTIFIER_AUTHORITY_URI = "https://www.opm.gov/"
OPM_WORKFORCE_DATA_URL = "https://data.opm.gov/explore-data/data/data-downloads"
OPM_EHRI_DATA_STANDARDS_URL = "https://data.opm.gov/data-standards/ehri-data-standards"
OPM_PLUM_DATA_URL = "https://www.opm.gov/about-us/open-government/plum-reporting/plum-data/"
OPM_PLUM_ALL_DATA_URL = "https://escs.opm.gov/escs-net/api/pbpub/download-data"

ResourceName = Literal[
    "payPlanCodes",
    "workScheduleCodes",
    "appointmentTypeCodes",
    "occupationalSeriesCodes",
    "plumPositionStatusCodes",
]
ResourceUse = Literal["deterministicMetadata"]

_DIGEST = re.compile(r"^sha256:([0-9a-f]{64})$")
_ALLOWED_HOSTS = frozenset({"data.opm.gov", "www.opm.gov"})
_RECORD_FIELDS = frozenset({"code", "label", "category"})

# Category names are the only vocabulary this module recognizes for the
# "category" field of a captured record; each maps to a code-shape rule and
# a ControlledIdentifier.kind. Codes outside this shape are refused; codes
# inside the shape but outside the pinned sample are accepted as unknown,
# never rejected, because none of these lists is asserted to be exhaustive.
_CATEGORY_PATTERNS: dict[str, re.Pattern[str]] = {
    "payPlan": re.compile(r"^[A-Z]{2}$"),
    "workSchedule": re.compile(r"^[A-Z]$"),
    "appointmentType": re.compile(r"^[0-9]{2}$"),
    "occupationalSeries": re.compile(r"^[0-9]{4}$"),
    "plumAppointmentAuthority": re.compile(r"^[A-Z]{2,6}$"),
    "plumIncumbentStatusMarker": re.compile(r"^[A-Z]{4,10}$"),
}
_CATEGORY_IDENTIFIER_KIND: dict[str, str] = {
    "payPlan": "opmPayPlanCode",
    "workSchedule": "opmWorkScheduleCode",
    "appointmentType": "opmAppointmentTypeCode",
    "occupationalSeries": "opmOccupationalSeriesCode",
    "plumAppointmentAuthority": "plumAppointmentAuthorityCode",
    "plumIncumbentStatusMarker": "plumIncumbentStatusMarker",
}


class OPMResourceError(ValueError):
    """Base class for OPM/PLUM controlled-code failures."""


class OPMAcquisitionError(OPMResourceError):
    """Exact official source bytes could not be acquired safely."""


class OPMSourceDriftError(OPMResourceError):
    """An OPM/PLUM source no longer matches the reviewed structure or pin."""


class OPMAssignmentError(OPMResourceError):
    """A workforce or PLUM record carries an unknown-shape or ungoverned code."""


@dataclass(frozen=True, slots=True)
class OPMConstantSource:
    """One official OPM or PLUM code-list artifact."""

    resource_name: ResourceName
    source_url: str
    filename: str
    expected_count: int
    categories: tuple[str, ...]
    # None of these lists was acquired from a verified live response; this
    # flag exists so a future pass can flip it once a real capture is pinned.
    is_closed_enumeration: bool

    def __post_init__(self) -> None:
        parsed = urlsplit(self.source_url)
        if parsed.scheme != "https" or parsed.hostname not in _ALLOWED_HOSTS:
            raise OPMAcquisitionError("source_url must be an official HTTPS opm.gov URL")
        if parsed.username is not None or parsed.password is not None:
            raise OPMAcquisitionError("source_url must not contain credentials")
        if not self.filename or Path(self.filename).name != self.filename:
            raise OPMAcquisitionError("filename must be one plain path component")
        if self.expected_count <= 0:
            raise OPMAcquisitionError("expected_count must be positive")
        if not self.categories or any(category not in _CATEGORY_PATTERNS for category in self.categories):
            raise OPMAcquisitionError("categories must be non-empty and recognized")
        if len(set(self.categories)) != len(self.categories):
            raise OPMAcquisitionError("categories must not repeat")


OPM_PAY_PLAN_CODES = OPMConstantSource(
    resource_name="payPlanCodes",
    source_url=OPM_WORKFORCE_DATA_URL,
    filename="opm-pay-plan-codes.json",
    expected_count=6,
    categories=("payPlan",),
    is_closed_enumeration=False,
)
OPM_WORK_SCHEDULE_CODES = OPMConstantSource(
    resource_name="workScheduleCodes",
    source_url=OPM_WORKFORCE_DATA_URL,
    filename="opm-work-schedule-codes.json",
    expected_count=5,
    categories=("workSchedule",),
    is_closed_enumeration=False,
)
OPM_APPOINTMENT_TYPE_CODES = OPMConstantSource(
    resource_name="appointmentTypeCodes",
    source_url=OPM_WORKFORCE_DATA_URL,
    filename="opm-appointment-type-codes.json",
    expected_count=4,
    categories=("appointmentType",),
    is_closed_enumeration=False,
)
OPM_OCCUPATIONAL_SERIES_CODES = OPMConstantSource(
    resource_name="occupationalSeriesCodes",
    source_url=OPM_WORKFORCE_DATA_URL,
    filename="opm-occupational-series-codes.json",
    expected_count=6,
    categories=("occupationalSeries",),
    is_closed_enumeration=False,
)
OPM_PLUM_POSITION_STATUS_CODES = OPMConstantSource(
    resource_name="plumPositionStatusCodes",
    source_url=OPM_PLUM_DATA_URL,
    filename="opm-plum-position-status-codes.json",
    expected_count=7,
    categories=("plumAppointmentAuthority", "plumIncumbentStatusMarker"),
    is_closed_enumeration=False,
)


@dataclass(frozen=True, slots=True)
class OPMSnapshotPin:
    """Exact identity of one official OPM or PLUM code-list capture."""

    source: OPMConstantSource
    retrieved_at: str
    expected_sha256: str
    expected_byte_length: int
    release_vintage: str | None = None
    requires_certification: bool = False

    def __post_init__(self) -> None:
        if _DIGEST.fullmatch(self.expected_sha256) is None:
            raise OPMAcquisitionError("expected_sha256 must be a lowercase sha256:<64 hex> digest")
        if self.expected_byte_length <= 0:
            raise OPMAcquisitionError("expected_byte_length must be positive")
        if not self.retrieved_at:
            raise OPMAcquisitionError("retrieved_at must not be empty")
        if self.release_vintage is not None and not self.release_vintage.strip():
            raise OPMAcquisitionError("release_vintage must not be blank when provided")


# These pins target a documented-shape sample fixture captured on 2026-08-03,
# not a verified live OPM/PLUM response (see module docstring). Replace with
# a live-fetched pin once a real capture has been reviewed.
OPM_PAY_PLAN_CODES_2026_08_03 = OPMSnapshotPin(
    source=OPM_PAY_PLAN_CODES,
    retrieved_at="2026-08-03T00:00:00Z",
    expected_sha256="sha256:925adfafa20e77218243fdaf117ea536c5bec08633999ad0faf716539c2b793a",
    expected_byte_length=487,
)
OPM_WORK_SCHEDULE_CODES_2026_08_03 = OPMSnapshotPin(
    source=OPM_WORK_SCHEDULE_CODES,
    retrieved_at="2026-08-03T00:00:00Z",
    expected_sha256="sha256:2a36a94088ae45dba51473852d721c8e13a32f555425df9f1d427370a544543f",
    expected_byte_length=362,
)
OPM_APPOINTMENT_TYPE_CODES_2026_08_03 = OPMSnapshotPin(
    source=OPM_APPOINTMENT_TYPE_CODES,
    retrieved_at="2026-08-03T00:00:00Z",
    expected_sha256="sha256:8ff37e93486e0a3aeed5aca915e718c45cbb16f62b80904b62d817b6fe2e85bd",
    expected_byte_length=377,
)
OPM_OCCUPATIONAL_SERIES_CODES_2026_08_03 = OPMSnapshotPin(
    source=OPM_OCCUPATIONAL_SERIES_CODES,
    retrieved_at="2026-08-03T00:00:00Z",
    expected_sha256="sha256:630568b89921ee9969582b237901a23077c1cf22c3f5d913dd1652eb6cfc9117",
    expected_byte_length=562,
)
OPM_PLUM_POSITION_STATUS_CODES_2026_08_03 = OPMSnapshotPin(
    source=OPM_PLUM_POSITION_STATUS_CODES,
    retrieved_at="2026-08-03T00:00:00Z",
    expected_sha256="sha256:9b3a8653da43402b26bdb361fcd4ed63c0d1c0d56430a4392f1e6d603e02bf8a",
    expected_byte_length=830,
    # The next PLUM Act release after the pinned sample; update once a real
    # edition's vintage is confirmed and its codes are re-pinned from bytes.
    release_vintage="2025",
    requires_certification=True,
)


@dataclass(frozen=True, slots=True)
class FetchedOPMResponse:
    """Provider-independent response returned by an injected fetcher."""

    body: bytes
    status_code: int
    content_type: str
    resolved_url: str


class OPMFetcher(Protocol):
    """Small transport boundary for official OPM/PLUM downloads."""

    def fetch(self, source_url: str, *, timeout_seconds: float) -> FetchedOPMResponse:
        """Fetch one response while preserving its exact body bytes."""


@dataclass(frozen=True, slots=True)
class AcquiredOPMSource:
    """One verified source object in the content-addressed store."""

    pin: OPMSnapshotPin
    path: Path
    sha256: str
    byte_length: int
    source_url: str
    resolved_url: str | None
    content_type: str
    acquisition_mode: AcquisitionMode
    cache_hit: bool
    local_source_path: Path | None


@dataclass(frozen=True, slots=True)
class OPMCode:
    """One exact publisher code, label, and category."""

    resource_name: ResourceName
    category: str
    use: ResourceUse
    publisher_label: str
    source_url: str
    identifiers: tuple[ControlledIdentifier, ...]
    is_general_subject_concept: bool = False


@dataclass(frozen=True, slots=True)
class ParsedOPMResource:
    """A parsed, digest-pinned OPM or PLUM code list."""

    source: OPMConstantSource
    retrieved_at: str
    source_sha256: str
    source_byte_length: int
    release_vintage: str | None
    requires_certification: bool
    codes: tuple[OPMCode, ...]
    gaps: tuple[str, ...]

    def by_code(self) -> dict[str, OPMCode]:
        """Index every row by its exact publisher code."""

        result: dict[str, OPMCode] = {}
        for entry in self.codes:
            kind = _CATEGORY_IDENTIFIER_KIND[entry.category]
            matches = [identifier for identifier in entry.identifiers if identifier.kind == kind]
            if len(matches) != 1:
                raise OPMSourceDriftError(f"{self.source.resource_name} row must retain exactly one {kind}")
            result[matches[0].value] = entry
        return result


OPM_PORTFOLIO_GAPS = (
    (
        "Neither data.opm.gov nor the PLUM data page publishes a stable per-resource "
        "JSON endpoint the way the LDA API does; source_url pins the known landing "
        "page and filename plus digest pin the specific captured artifact."
    ),
    (
        "None of these code lists was acquired from a live OPM response for this "
        "pass; every pinned sample is a documented-shape placeholder pending a "
        "verified live capture, and expected_count deliberately rejects a "
        "differently sized real capture until it is reviewed and re-pinned."
    ),
    (
        "The occupational series code list is not exhaustive; occupational series "
        "codes are validated by 4-digit shape only, and an unmatched code is never "
        "treated as invalid."
    ),
    (
        "PLUM Act position data requires agency certification and a pinned release "
        "vintage before its appointment-authority and incumbent-status codes are "
        "treated as workforce metadata; incumbent identity for protected positions "
        "may be redacted, and a redacted record must carry a redaction_reason."
    ),
)


def sha256_digest(payload: bytes) -> str:
    """Return the canonical RefSpec SHA-256 spelling."""

    return "sha256:" + hashlib.sha256(payload).hexdigest()


@dataclass(frozen=True, slots=True)
class OPMEHRIDataElement:
    """One data-element row from OPM's official EHRI workbook."""

    name: str
    description: str
    data_format: str
    data_length: str
    valid_values: str
    current_values: str
    past_values: str


@dataclass(frozen=True, slots=True)
class OPMEHRIValue:
    """One current or past publisher value from OPM's official EHRI workbook."""

    name: str
    code: str
    explanation: str
    from_date: str
    through_date: str


@dataclass(frozen=True, slots=True)
class OPMEHRIDataStandardsExport:
    """The complete three-sheet EHRI data-standards export."""

    source_sha256: str
    source_byte_length: int
    fields: tuple[OPMEHRIDataElement, ...]
    current_values: tuple[OPMEHRIValue, ...]
    past_values: tuple[OPMEHRIValue, ...]

    def current_values_for(self, name: str) -> tuple[OPMEHRIValue, ...]:
        """Return all current publisher values for one exact EHRI element name."""

        return tuple(value for value in self.current_values if value.name == name)


@dataclass(frozen=True, slots=True)
class OPMPLUMRow:
    """One exact row from OPM's PLUM all-data CSV."""

    agency_name: str
    organization_name: str
    position_title: str
    position_status: str
    appointment_type: str
    expiration_date: str
    level_grade_pay: str
    location: str
    incumbent_first_name: str
    incumbent_last_name: str
    pay_plan: str
    tenure: str
    incumbent_begin_date: str
    incumbent_vacate_date: str


@dataclass(frozen=True, slots=True)
class OPMPLUMAllDataExport:
    """One complete official PLUM CSV capture and its observed code values."""

    source_sha256: str
    source_byte_length: int
    records: tuple[OPMPLUMRow, ...]
    appointment_types: tuple[str, ...]
    position_statuses: tuple[str, ...]
    pay_plans: tuple[str, ...]


_EHRI_SHEETS = ("AllDataElements", "CurrentValues", "PastValues")
_EHRI_ELEMENT_HEADER = (
    "Name",
    "Description",
    "Data Format",
    "Data Length",
    "Valid Values",
    "Current Values",
    "Past Values",
)
_EHRI_VALUE_HEADER = ("Name", "Code", "Explanation", "From Date", "Through Date")
_PLUM_HEADER = (
    "AgencyName",
    "OrganizationName",
    "PositionTitle",
    "PositionStatus",
    "AppointmentTypeDescription",
    "ExpirationDate",
    "LevelGradePay",
    "Location",
    "IncumbentFirstName",
    "IncumbentLastName",
    "PaymentPlanDescription",
    "Tenure",
    "IncumbentBeginDate",
    "IncumbentVacateDate",
)


def _cell_text(value: object) -> str:
    return "" if value is None else str(value)


def parse_opm_ehri_data_standards_xlsx(payload: bytes) -> OPMEHRIDataStandardsExport:
    """Parse all fields and all current/past values from the official OPM workbook."""

    if not payload:
        raise OPMSourceDriftError("EHRI data-standards workbook is empty")
    try:
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    except Exception as error:
        raise OPMSourceDriftError("EHRI data-standards source is not a readable XLSX workbook") from error
    if tuple(workbook.sheetnames) != _EHRI_SHEETS:
        raise OPMSourceDriftError(f"EHRI workbook sheets drifted: {workbook.sheetnames!r}")

    element_sheet = workbook[_EHRI_SHEETS[0]]
    element_rows = element_sheet.iter_rows(values_only=True)
    if tuple(_cell_text(value) for value in next(element_rows)) != _EHRI_ELEMENT_HEADER:
        raise OPMSourceDriftError("EHRI AllDataElements header drifted")
    fields = tuple(
        OPMEHRIDataElement(*(_cell_text(value) for value in row))
        for row in element_rows
        if any(value is not None and str(value) for value in row)
    )

    def values_from(sheet_name: str) -> tuple[OPMEHRIValue, ...]:
        rows = workbook[sheet_name].iter_rows(values_only=True)
        if tuple(_cell_text(value) for value in next(rows)) != _EHRI_VALUE_HEADER:
            raise OPMSourceDriftError(f"EHRI {sheet_name} header drifted")
        return tuple(
            OPMEHRIValue(*(_cell_text(value) for value in row))
            for row in rows
            if any(value is not None and str(value) for value in row)
        )

    current_values = values_from("CurrentValues")
    past_values = values_from("PastValues")
    if not fields or not current_values or not past_values:
        raise OPMSourceDriftError("EHRI workbook omitted a required non-empty data sheet")
    return OPMEHRIDataStandardsExport(
        source_sha256=sha256_digest(payload),
        source_byte_length=len(payload),
        fields=fields,
        current_values=current_values,
        past_values=past_values,
    )


def parse_opm_plum_all_data_csv(payload: bytes) -> OPMPLUMAllDataExport:
    """Parse every exact row and observed code value from OPM's PLUM CSV export."""

    try:
        text = payload.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise OPMSourceDriftError("PLUM all-data source is not UTF-8 CSV") from error
    reader = csv.DictReader(io.StringIO(text, newline=""))
    if tuple(reader.fieldnames or ()) != _PLUM_HEADER:
        raise OPMSourceDriftError(f"PLUM all-data header drifted: {reader.fieldnames!r}")
    records = tuple(
        OPMPLUMRow(
            agency_name=row["AgencyName"],
            organization_name=row["OrganizationName"],
            position_title=row["PositionTitle"],
            position_status=row["PositionStatus"],
            appointment_type=row["AppointmentTypeDescription"],
            expiration_date=row["ExpirationDate"],
            level_grade_pay=row["LevelGradePay"],
            location=row["Location"],
            incumbent_first_name=row["IncumbentFirstName"],
            incumbent_last_name=row["IncumbentLastName"],
            pay_plan=row["PaymentPlanDescription"],
            tenure=row["Tenure"],
            incumbent_begin_date=row["IncumbentBeginDate"],
            incumbent_vacate_date=row["IncumbentVacateDate"],
        )
        for row in reader
    )
    if not records:
        raise OPMSourceDriftError("PLUM all-data CSV contains no records")
    return OPMPLUMAllDataExport(
        source_sha256=sha256_digest(payload),
        source_byte_length=len(payload),
        records=records,
        appointment_types=tuple(sorted({row.appointment_type for row in records})),
        position_statuses=tuple(sorted({row.position_status for row in records})),
        pay_plans=tuple(sorted({row.pay_plan for row in records})),
    )


def _validate_resolved_url(value: str) -> None:
    parsed = urlsplit(value)
    if parsed.scheme != "https" or parsed.hostname not in _ALLOWED_HOSTS:
        raise OPMAcquisitionError("fetcher resolved_url must remain on an official HTTPS opm.gov host")
    if parsed.username is not None or parsed.password is not None:
        raise OPMAcquisitionError("fetcher resolved_url must not contain credentials")


def _verify_payload(payload: bytes, pin: OPMSnapshotPin, *, location: str) -> tuple[str, int]:
    byte_length = len(payload)
    if byte_length != pin.expected_byte_length:
        raise OPMSourceDriftError(
            f"{location} byte length drift: expected {pin.expected_byte_length}, got {byte_length}"
        )
    actual_sha256 = sha256_digest(payload)
    if actual_sha256 != pin.expected_sha256:
        raise OPMSourceDriftError(f"{location} digest drift: expected {pin.expected_sha256}, got {actual_sha256}")
    try:
        json.loads(payload)
    except (UnicodeDecodeError, json.JSONDecodeError) as error:
        raise OPMSourceDriftError(f"{location} is not valid JSON") from error
    return actual_sha256, byte_length


def _verify_existing(path: Path, pin: OPMSnapshotPin) -> AcquiredOPMSource:
    if path.is_symlink() or not path.is_file():
        raise OPMAcquisitionError(f"content-addressed target is not a regular file: {path}")
    actual_sha256, byte_length = _verify_payload(
        path.read_bytes(),
        pin,
        location="cached OPM source",
    )
    return AcquiredOPMSource(
        pin=pin,
        path=path,
        sha256=actual_sha256,
        byte_length=byte_length,
        source_url=pin.source.source_url,
        resolved_url=None,
        content_type="application/json",
        acquisition_mode="cache",
        cache_hit=True,
        local_source_path=None,
    )


def _publish_payload(
    payload: bytes,
    pin: OPMSnapshotPin,
    final_path: Path,
    *,
    content_type: str,
    acquisition_mode: Literal["local", "fetcher"],
    resolved_url: str | None,
    local_source_path: Path | None,
) -> AcquiredOPMSource:
    actual_sha256, byte_length = _verify_payload(
        payload,
        pin,
        location=f"{acquisition_mode} OPM source",
    )
    final_path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=".acquire-",
        suffix=".tmp",
        dir=final_path.parent,
    )
    temporary_path = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as output:
            descriptor = -1
            output.write(payload)
            output.flush()
            os.fsync(output.fileno())
        try:
            os.link(temporary_path, final_path)
        except FileExistsError:
            return _verify_existing(final_path, pin)
        return AcquiredOPMSource(
            pin=pin,
            path=final_path,
            sha256=actual_sha256,
            byte_length=byte_length,
            source_url=pin.source.source_url,
            resolved_url=resolved_url,
            content_type=content_type,
            acquisition_mode=acquisition_mode,
            cache_hit=False,
            local_source_path=local_source_path,
        )
    finally:
        if descriptor >= 0:
            os.close(descriptor)
        temporary_path.unlink(missing_ok=True)


def acquire_opm_constants(
    pin: OPMSnapshotPin,
    store_dir: Path,
    *,
    source_path: Path | None = None,
    fetcher: OPMFetcher | None = None,
    timeout_seconds: float = 30.0,
) -> AcquiredOPMSource:
    """Acquire one exact OPM/PLUM code-list capture through a provider-neutral boundary."""

    if timeout_seconds <= 0:
        raise OPMAcquisitionError("timeout_seconds must be positive")
    if source_path is not None and fetcher is not None:
        raise OPMAcquisitionError("provide source_path or fetcher, not both")
    digest_hex = cast(re.Match[str], _DIGEST.fullmatch(pin.expected_sha256)).group(1)
    final_path = Path(store_dir) / "sha256" / digest_hex / pin.source.filename
    if final_path.exists() or final_path.is_symlink():
        return _verify_existing(final_path, pin)

    if source_path is not None:
        local_path = Path(source_path)
        if local_path.is_symlink() or not local_path.is_file():
            raise OPMAcquisitionError(f"local OPM source is not a regular file: {local_path}")
        return _publish_payload(
            local_path.read_bytes(),
            pin,
            final_path,
            content_type="application/json",
            acquisition_mode="local",
            resolved_url=None,
            local_source_path=local_path.resolve(),
        )

    if fetcher is None:
        raise OPMAcquisitionError("OPM/PLUM source is not cached; provide source_path or an injected fetcher")
    fetched = fetcher.fetch(pin.source.source_url, timeout_seconds=timeout_seconds)
    if fetched.status_code != 200:
        raise OPMAcquisitionError(f"could not acquire {pin.source.source_url}: HTTP {fetched.status_code}")
    _validate_resolved_url(fetched.resolved_url)
    media_type = fetched.content_type.partition(";")[0].strip().lower()
    if media_type != "application/json":
        raise OPMSourceDriftError(f"OPM/PLUM source content type drifted to {fetched.content_type!r}")
    return _publish_payload(
        fetched.body,
        pin,
        final_path,
        content_type=fetched.content_type,
        acquisition_mode="fetcher",
        resolved_url=fetched.resolved_url,
        local_source_path=None,
    )


def parse_opm_constants(acquired: AcquiredOPMSource) -> ParsedOPMResource:
    """Parse exact ``code``/``label``/``category`` records without inventing labels."""

    payload = acquired.path.read_bytes()
    _verify_payload(payload, acquired.pin, location="parsed OPM source")
    try:
        root = json.loads(payload)
    except (UnicodeDecodeError, json.JSONDecodeError) as error:
        raise OPMSourceDriftError("OPM/PLUM source payload is not valid JSON") from error
    if not isinstance(root, list):
        raise OPMSourceDriftError("OPM/PLUM source payload must be an array")
    if len(root) != acquired.pin.source.expected_count:
        raise OPMSourceDriftError(
            f"{acquired.pin.source.resource_name} count drift: expected "
            f"{acquired.pin.source.expected_count}, parsed {len(root)}"
        )

    allowed_categories = acquired.pin.source.categories
    parsed: list[OPMCode] = []
    for ordinal, record in enumerate(root, start=1):
        if not isinstance(record, Mapping):
            raise OPMSourceDriftError(f"OPM/PLUM record {ordinal} must be an object")
        if set(record) != _RECORD_FIELDS:
            raise OPMSourceDriftError(f"OPM/PLUM record {ordinal} fields drifted: {sorted(record)}")
        category = record["category"]
        if category not in allowed_categories:
            raise OPMSourceDriftError(f"OPM/PLUM record {ordinal} has unexpected category {category!r}")
        code = record["code"]
        label = record["label"]
        pattern = _CATEGORY_PATTERNS[category]
        if not isinstance(code, str) or pattern.fullmatch(code) is None:
            raise OPMSourceDriftError(f"OPM/PLUM record {ordinal} has malformed {category} code {code!r}")
        if not isinstance(label, str) or not label.strip() or label != label.strip():
            raise OPMSourceDriftError(f"OPM/PLUM record {ordinal} has malformed publisher label")
        identifier = ControlledIdentifier(
            value=code,
            kind=_CATEGORY_IDENTIFIER_KIND[category],
            authority_uri=OPM_IDENTIFIER_AUTHORITY_URI,
            source_uri=acquired.pin.source.source_url,
            observed_at=acquired.pin.retrieved_at,
            effective_at=None,
            source_digest=acquired.sha256,
        )
        parsed.append(
            OPMCode(
                resource_name=acquired.pin.source.resource_name,
                category=category,
                use="deterministicMetadata",
                publisher_label=label,
                source_url=acquired.pin.source.source_url,
                identifiers=(identifier,),
            )
        )
    code_values = {entry.identifiers[0].value for entry in parsed}
    if len(code_values) != len(parsed):
        raise OPMSourceDriftError("OPM/PLUM source contains duplicate publisher codes")

    return ParsedOPMResource(
        source=acquired.pin.source,
        retrieved_at=acquired.pin.retrieved_at,
        source_sha256=acquired.sha256,
        source_byte_length=acquired.byte_length,
        release_vintage=acquired.pin.release_vintage,
        requires_certification=acquired.pin.requires_certification,
        codes=tuple(parsed),
        gaps=OPM_PORTFOLIO_GAPS,
    )


@dataclass(frozen=True, slots=True)
class OPMControlPortfolio:
    """The five imported OPM/PLUM code resources and known gaps."""

    pay_plan_codes: ParsedOPMResource
    work_schedule_codes: ParsedOPMResource
    appointment_type_codes: ParsedOPMResource
    occupational_series_codes: ParsedOPMResource
    plum_position_status_codes: ParsedOPMResource
    gaps: tuple[str, ...]


_PORTFOLIO_RESOURCE_NAMES = frozenset(
    {
        "payPlanCodes",
        "workScheduleCodes",
        "appointmentTypeCodes",
        "occupationalSeriesCodes",
        "plumPositionStatusCodes",
    }
)


def assemble_opm_control_portfolio(
    resources: Sequence[ParsedOPMResource],
) -> OPMControlPortfolio:
    """Require exactly the five known OPM/PLUM resources and retain their gaps."""

    by_name = {resource.source.resource_name: resource for resource in resources}
    if len(resources) != 5 or set(by_name) != _PORTFOLIO_RESOURCE_NAMES:
        raise OPMSourceDriftError("OPM control portfolio requires exactly the five known workforce/PLUM resources")
    return OPMControlPortfolio(
        pay_plan_codes=by_name["payPlanCodes"],
        work_schedule_codes=by_name["workScheduleCodes"],
        appointment_type_codes=by_name["appointmentTypeCodes"],
        occupational_series_codes=by_name["occupationalSeriesCodes"],
        plum_position_status_codes=by_name["plumPositionStatusCodes"],
        gaps=OPM_PORTFOLIO_GAPS,
    )


@dataclass(frozen=True, slots=True)
class OPMFieldAssignment:
    """One workforce or PLUM field validated against its pinned code shape."""

    source_field: str
    code: str
    publisher_label: str | None
    identifiers: tuple[ControlledIdentifier, ...]
    # False means the code is shape-valid but absent from the pinned sample;
    # it is accepted, not rejected, because the sample is not exhaustive.
    in_pinned_sample: bool


def _lookup_field(
    raw_value: object,
    *,
    source_field: str,
    resource: ParsedOPMResource,
    category: str,
) -> OPMFieldAssignment:
    if not isinstance(raw_value, str) or not raw_value:
        raise OPMAssignmentError(f"{source_field} must carry a non-empty string code")
    pattern = _CATEGORY_PATTERNS[category]
    if pattern.fullmatch(raw_value) is None:
        raise OPMAssignmentError(f"{source_field} value {raw_value!r} does not match the {category} code shape")
    match = resource.by_code().get(raw_value)
    if match is not None:
        return OPMFieldAssignment(
            source_field=source_field,
            code=raw_value,
            publisher_label=match.publisher_label,
            identifiers=match.identifiers,
            in_pinned_sample=True,
        )
    identifier = ControlledIdentifier(
        value=raw_value,
        kind=_CATEGORY_IDENTIFIER_KIND[category],
        authority_uri=OPM_IDENTIFIER_AUTHORITY_URI,
        source_uri=resource.source.source_url,
        observed_at=resource.retrieved_at,
        effective_at=None,
        source_digest=resource.source_sha256,
    )
    return OPMFieldAssignment(
        source_field=source_field,
        code=raw_value,
        publisher_label=None,
        identifiers=(identifier,),
        in_pinned_sample=False,
    )


@dataclass(frozen=True, slots=True)
class ValidatedOPMWorkforceCodes:
    """Code evidence retained from one workforce entity/observation record."""

    pay_plan: OPMFieldAssignment
    work_schedule: OPMFieldAssignment
    appointment_type: OPMFieldAssignment | None
    occupational_series: OPMFieldAssignment
    gaps: tuple[str, ...]


def validate_workforce_observation_codes(
    observation: Mapping[str, object],
    portfolio: OPMControlPortfolio,
) -> ValidatedOPMWorkforceCodes:
    """Validate the FedScope-style codes carried by one workforce record."""

    pay_plan = _lookup_field(
        observation.get("pay_plan"),
        source_field="pay_plan",
        resource=portfolio.pay_plan_codes,
        category="payPlan",
    )
    work_schedule = _lookup_field(
        observation.get("work_schedule"),
        source_field="work_schedule",
        resource=portfolio.work_schedule_codes,
        category="workSchedule",
    )
    occupational_series = _lookup_field(
        observation.get("occupational_series"),
        source_field="occupational_series",
        resource=portfolio.occupational_series_codes,
        category="occupationalSeries",
    )
    raw_appointment_type = observation.get("appointment_type")
    appointment_type = (
        None
        if raw_appointment_type is None
        else _lookup_field(
            raw_appointment_type,
            source_field="appointment_type",
            resource=portfolio.appointment_type_codes,
            category="appointmentType",
        )
    )
    return ValidatedOPMWorkforceCodes(
        pay_plan=pay_plan,
        work_schedule=work_schedule,
        appointment_type=appointment_type,
        occupational_series=occupational_series,
        gaps=portfolio.gaps,
    )


@dataclass(frozen=True, slots=True)
class ValidatedPLUMPositionCodes:
    """Code evidence retained from one certified PLUM position record."""

    appointment_authority: OPMFieldAssignment
    incumbent_status: Literal["named", "vacant", "redacted"]
    incumbent_status_marker: OPMFieldAssignment | None
    redaction_reason: str | None
    release_vintage: str
    gaps: tuple[str, ...]


def validate_plum_position_codes(
    record: Mapping[str, object],
    portfolio: OPMControlPortfolio,
) -> ValidatedPLUMPositionCodes:
    """Validate one PLUM record's codes, enforcing certification, vintage, and redaction rules."""

    resource = portfolio.plum_position_status_codes
    if record.get("release_certified") is not True:
        raise OPMAssignmentError(
            "a PLUM record must be marked release_certified before its codes are treated as workforce metadata"
        )
    raw_vintage = record.get("release_vintage")
    if not isinstance(raw_vintage, str) or raw_vintage != resource.release_vintage:
        raise OPMAssignmentError(
            f"PLUM record release_vintage {raw_vintage!r} does not match the pinned code list vintage "
            f"{resource.release_vintage!r}"
        )
    appointment_authority = _lookup_field(
        record.get("appointment_authority"),
        source_field="appointment_authority",
        resource=resource,
        category="plumAppointmentAuthority",
    )
    raw_status = record.get("incumbent_status")
    if raw_status not in {"named", "vacant", "redacted"}:
        raise OPMAssignmentError(f"unknown PLUM incumbent_status {raw_status!r}")
    incumbent_status = cast(Literal["named", "vacant", "redacted"], raw_status)

    marker: OPMFieldAssignment | None = None
    redaction_reason: str | None = None
    if incumbent_status == "vacant":
        marker = _lookup_field(
            "VACANT",
            source_field="incumbent_status",
            resource=resource,
            category="plumIncumbentStatusMarker",
        )
    elif incumbent_status == "redacted":
        raw_reason = record.get("redaction_reason")
        if not isinstance(raw_reason, str) or not raw_reason.strip():
            raise OPMAssignmentError("a redacted PLUM record must carry a non-empty redaction_reason")
        redaction_reason = raw_reason
        marker = _lookup_field(
            "REDACTED",
            source_field="incumbent_status",
            resource=resource,
            category="plumIncumbentStatusMarker",
        )

    return ValidatedPLUMPositionCodes(
        appointment_authority=appointment_authority,
        incumbent_status=incumbent_status,
        incumbent_status_marker=marker,
        redaction_reason=redaction_reason,
        release_vintage=resource.release_vintage or "",
        gaps=resource.gaps,
    )


# --- Deterministic closed packages -----------------------------------------
#
# The functions and dataclasses below package one parsed OPM/PLUM resource as
# a development-only ``controlledCodeList`` resource using the shared
# source-controlled-resource format, mirroring how other RefSpec controlled
# code lists are packaged. Building never claims concept identity, and the
# reader below only reopens a package that reproduces its bytes exactly.

OPM_CONTROLLED_LIST_PACKAGE_VERSION = "opm-controlled-list-package-v1"
_OBSERVATION_FIELDS = frozenset(
    {
        "id",
        "sourceArtifact",
        "sourcePath",
        "sourceOrdinal",
        "labels",
        "identifiers",
        "uses",
        "conceptIdentityClaimed",
        "category",
    }
)


class OPMControlledListPackageError(OPMResourceError):
    """An OPM/PLUM package differs from its exact source or declared use."""


@dataclass(frozen=True, slots=True)
class OPMControlledListPackageSpec:
    """Pinned identity and use of one OPM/PLUM controlled-list package."""

    resource_name: ResourceName
    resource_id: str
    title: str
    pin: OPMSnapshotPin
    known_gaps: tuple[Mapping[str, str], ...]
    expected_logical_digest: str

    def __post_init__(self) -> None:
        if self.resource_name != self.pin.source.resource_name:
            raise OPMControlledListPackageError("package resource_name differs from its source pin")
        if not self.resource_id or not self.title:
            raise OPMControlledListPackageError("package identity fields must not be empty")
        if _DIGEST.fullmatch(self.expected_logical_digest) is None:
            raise OPMControlledListPackageError("expected_logical_digest must be a SHA-256 digest")


_UNVERIFIED_SAMPLE_GAP = MappingProxyType(
    {
        "kind": "unverifiedSampleCapture",
        "reason": OPM_PORTFOLIO_GAPS[1],
    }
)
_NO_STABLE_ENDPOINT_GAP = MappingProxyType(
    {
        "kind": "noStablePerResourceEndpoint",
        "reason": OPM_PORTFOLIO_GAPS[0],
    }
)
_NOT_EXHAUSTIVE_SERIES_GAP = MappingProxyType(
    {
        "kind": "occupationalSeriesSampleNotExhaustive",
        "reason": OPM_PORTFOLIO_GAPS[2],
    }
)
_PLUM_RULES_GAP = MappingProxyType(
    {
        "kind": "plumCertificationRedactionVintageRules",
        "reason": OPM_PORTFOLIO_GAPS[3],
    }
)

OPM_PAY_PLAN_CODE_PACKAGE = OPMControlledListPackageSpec(
    resource_name="payPlanCodes",
    resource_id="opm-pay-plan-codes-2026-08-03",
    title="OPM Pay Plan Codes, captured 2026-08-03",
    pin=OPM_PAY_PLAN_CODES_2026_08_03,
    known_gaps=(_NO_STABLE_ENDPOINT_GAP, _UNVERIFIED_SAMPLE_GAP),
    expected_logical_digest="sha256:6f87f3c43253b1e4bd5710b52db5502e3566be278f88df1423ecd08d0ecb03c6",
)
OPM_WORK_SCHEDULE_CODE_PACKAGE = OPMControlledListPackageSpec(
    resource_name="workScheduleCodes",
    resource_id="opm-work-schedule-codes-2026-08-03",
    title="OPM Work Schedule Codes, captured 2026-08-03",
    pin=OPM_WORK_SCHEDULE_CODES_2026_08_03,
    known_gaps=(_NO_STABLE_ENDPOINT_GAP, _UNVERIFIED_SAMPLE_GAP),
    expected_logical_digest="sha256:4f494567ad875e707d9d6fcad40b68a044c060fbcc3df73d6e35e0edaf7a7625",
)
OPM_APPOINTMENT_TYPE_CODE_PACKAGE = OPMControlledListPackageSpec(
    resource_name="appointmentTypeCodes",
    resource_id="opm-appointment-type-codes-2026-08-03",
    title="OPM Appointment Type Codes, captured 2026-08-03",
    pin=OPM_APPOINTMENT_TYPE_CODES_2026_08_03,
    known_gaps=(_NO_STABLE_ENDPOINT_GAP, _UNVERIFIED_SAMPLE_GAP),
    expected_logical_digest="sha256:bf164a6e12e37da7a183c76da8bbbb49cd95aaaad42d2fe84a63d4abd31b5233",
)
OPM_OCCUPATIONAL_SERIES_CODE_PACKAGE = OPMControlledListPackageSpec(
    resource_name="occupationalSeriesCodes",
    resource_id="opm-occupational-series-codes-2026-08-03",
    title="OPM Occupational Series Codes, captured 2026-08-03",
    pin=OPM_OCCUPATIONAL_SERIES_CODES_2026_08_03,
    known_gaps=(_NO_STABLE_ENDPOINT_GAP, _UNVERIFIED_SAMPLE_GAP, _NOT_EXHAUSTIVE_SERIES_GAP),
    expected_logical_digest="sha256:342e489e28760277919743a7e7f29d5bc4363ff061e4be076f6ad992ecd3c269",
)
OPM_PLUM_POSITION_STATUS_CODE_PACKAGE = OPMControlledListPackageSpec(
    resource_name="plumPositionStatusCodes",
    resource_id="opm-plum-position-status-codes-2026-08-03",
    title="PLUM Position Appointment-Authority and Status Codes, captured 2026-08-03",
    pin=OPM_PLUM_POSITION_STATUS_CODES_2026_08_03,
    known_gaps=(_NO_STABLE_ENDPOINT_GAP, _UNVERIFIED_SAMPLE_GAP, _PLUM_RULES_GAP),
    expected_logical_digest="sha256:5c0bd5777351447e84be18719cf8fa1065d19c305aa8c930aa5a7dec55d792a7",
)
OPM_CONTROLLED_LIST_PACKAGES = (
    OPM_PAY_PLAN_CODE_PACKAGE,
    OPM_WORK_SCHEDULE_CODE_PACKAGE,
    OPM_APPOINTMENT_TYPE_CODE_PACKAGE,
    OPM_OCCUPATIONAL_SERIES_CODE_PACKAGE,
    OPM_PLUM_POSITION_STATUS_CODE_PACKAGE,
)
_PACKAGE_BY_RESOURCE_ID = MappingProxyType({spec.resource_id: spec for spec in OPM_CONTROLLED_LIST_PACKAGES})


def _sha256(payload: bytes) -> str:
    return "sha256:" + hashlib.sha256(payload).hexdigest()


def _parse_exact_source(
    spec: OPMControlledListPackageSpec,
    payload: bytes,
) -> ParsedOPMResource:
    with tempfile.TemporaryDirectory(prefix="refspec-opm-package-") as temporary:
        root = Path(temporary)
        source_path = root / spec.pin.source.filename
        source_path.write_bytes(payload)
        acquired = acquire_opm_constants(
            spec.pin,
            root / "store",
            source_path=source_path,
        )
        return parse_opm_constants(acquired)


def _identifier_payload(
    *,
    identifier: ControlledIdentifier,
    source_path: str,
) -> dict[str, Any]:
    return {
        "value": identifier.value,
        "kind": identifier.kind,
        "authorityUri": identifier.authority_uri,
        "sourceUri": identifier.source_uri,
        "sourcePath": f"{source_path}.code",
        "observedAt": identifier.observed_at,
        "sourceDigest": identifier.source_digest,
    }


def _observation_id(
    *,
    spec: OPMControlledListPackageSpec,
    source_path: str,
    identifiers: Sequence[Mapping[str, Any]],
) -> str:
    identity = {
        "packageVersion": OPM_CONTROLLED_LIST_PACKAGE_VERSION,
        "resourceId": spec.resource_id,
        "sourceArtifact": spec.pin.source.source_url,
        "sourcePath": source_path,
        "identifiers": [
            {
                "value": identifier["value"],
                "kind": identifier["kind"],
                "authorityUri": identifier["authorityUri"],
            }
            for identifier in identifiers
        ],
    }
    digest = hashlib.sha256(canonical_json(identity).encode("utf-8")).hexdigest()
    return f"urn:ref:source-observation:{spec.resource_id}:{digest}"


def _observations(
    spec: OPMControlledListPackageSpec,
    resource: ParsedOPMResource,
) -> tuple[Mapping[str, Any], ...]:
    if resource.source != spec.pin.source:
        raise OPMControlledListPackageError("parsed resource differs from its package source")
    if resource.source_sha256 != spec.pin.expected_sha256:
        raise OPMControlledListPackageError("parsed resource digest differs from its package source")
    if len(resource.codes) != spec.pin.source.expected_count:
        raise OPMControlledListPackageError("parsed resource count differs from its package source")

    result: list[Mapping[str, Any]] = []
    for ordinal, code in enumerate(resource.codes):
        if code.resource_name != spec.resource_name or code.is_general_subject_concept:
            raise OPMControlledListPackageError(f"{spec.resource_name} row {ordinal} has an incompatible type")
        source_path = f"$[{ordinal}]"
        identifiers = tuple(
            _identifier_payload(identifier=identifier, source_path=source_path) for identifier in code.identifiers
        )
        result.append(
            {
                "id": _observation_id(spec=spec, source_path=source_path, identifiers=identifiers),
                "sourceArtifact": spec.pin.source.source_url,
                "sourcePath": source_path,
                # This ordinal is a source locator only, never derived identity.
                "sourceOrdinal": ordinal,
                "labels": [
                    {
                        "value": code.publisher_label,
                        "language": "en",
                        "role": "preferred",
                    }
                ],
                "identifiers": list(identifiers),
                "uses": [cast(PackageResourceUse, "deterministicMetadata")],
                "conceptIdentityClaimed": False,
                "category": code.category,
            }
        )
    return tuple(result)


def build_opm_controlled_list_package(
    spec: OPMControlledListPackageSpec,
    source_path: Path,
) -> SourceControlledResourceBundle:
    """Build one exact, development-only OPM/PLUM controlled-list package."""

    path = Path(source_path)
    if path.is_symlink() or not path.is_file():
        raise OPMControlledListPackageError(f"OPM/PLUM controlled-list source is not a regular file: {path}")
    payload = path.read_bytes()
    resource = _parse_exact_source(spec, payload)
    return build_source_controlled_resource_bundle(
        resource_id=spec.resource_id,
        title=spec.title,
        resource_kind="controlledCodeList",
        identity_status="publisherIdentifiersPreserved",
        uses=(cast(PackageResourceUse, "deterministicMetadata"),),
        captured_at=spec.pin.retrieved_at,
        observations=_observations(spec, resource),
        source_artifacts={spec.pin.source.source_url: payload},
        source_observed_count=spec.pin.source.expected_count,
        gaps=spec.known_gaps,
    )


@dataclass(frozen=True, slots=True)
class OPMControlledListView:
    """An OPM/PLUM package reopened against its external pin and source rules."""

    package: SourceControlledResourceView
    spec: OPMControlledListPackageSpec
    observations_by_code: Mapping[str, Mapping[str, Any]]

    @classmethod
    def open(cls, path: Path) -> OPMControlledListView:
        """Open one known OPM/PLUM package and rebuild it from retained source bytes."""

        package = SourceControlledResourceView.open(path)
        resource_id = package.resource_manifest.get("resourceId")
        if not isinstance(resource_id, str) or resource_id not in _PACKAGE_BY_RESOURCE_ID:
            raise OPMControlledListPackageError(f"unknown OPM/PLUM controlled-list resource {resource_id!r}")
        spec = _PACKAGE_BY_RESOURCE_ID[resource_id]
        if package.logical_digest != spec.expected_logical_digest:
            raise OPMControlledListPackageError(f"{resource_id} logical digest differs from its external pin")
        source_bytes = package.source_artifact_bytes(spec.pin.source.source_url)
        if len(source_bytes) != spec.pin.expected_byte_length or _sha256(source_bytes) != spec.pin.expected_sha256:
            raise OPMControlledListPackageError(f"{resource_id} retained source differs from its dated pin")
        rebuilt = build_source_controlled_resource_bundle(
            resource_id=spec.resource_id,
            title=spec.title,
            resource_kind="controlledCodeList",
            identity_status="publisherIdentifiersPreserved",
            uses=(cast(PackageResourceUse, "deterministicMetadata"),),
            captured_at=spec.pin.retrieved_at,
            observations=_observations(spec, _parse_exact_source(spec, source_bytes)),
            source_artifacts={spec.pin.source.source_url: source_bytes},
            source_observed_count=spec.pin.source.expected_count,
            gaps=spec.known_gaps,
        )
        if rebuilt.artifact_bytes() != {
            relative_path: (Path(path) / relative_path).read_bytes() for relative_path in rebuilt.artifact_bytes()
        }:
            raise OPMControlledListPackageError(f"{resource_id} package differs from its deterministic OPM build")

        by_code: dict[str, Mapping[str, Any]] = {}
        for ordinal, observation in enumerate(package.observations):
            if set(observation) != _OBSERVATION_FIELDS:
                raise OPMControlledListPackageError(f"{resource_id} observation {ordinal} has unexpected fields")
            if len(observation["identifiers"]) != 1:
                raise OPMControlledListPackageError(f"{resource_id} observation {ordinal} lacks one publisher code")
            code = observation["identifiers"][0]["value"]
            if code in by_code:
                raise OPMControlledListPackageError(f"{resource_id} repeats publisher code {code!r}")
            by_code[code] = observation
        return cls(
            package=package,
            spec=spec,
            observations_by_code=MappingProxyType(by_code),
        )

    def lookup_code(self, value: str) -> Mapping[str, Any] | None:
        """Return one exact source observation by publisher code."""

        return self.observations_by_code.get(value)


__all__ = [
    "OPM_APPOINTMENT_TYPE_CODES",
    "OPM_APPOINTMENT_TYPE_CODES_2026_08_03",
    "OPM_APPOINTMENT_TYPE_CODE_PACKAGE",
    "OPM_CONTROLLED_LIST_PACKAGES",
    "OPM_CONTROLLED_LIST_PACKAGE_VERSION",
    "OPM_EHRI_DATA_STANDARDS_URL",
    "OPM_IDENTIFIER_AUTHORITY_URI",
    "OPM_OCCUPATIONAL_SERIES_CODES",
    "OPM_OCCUPATIONAL_SERIES_CODES_2026_08_03",
    "OPM_OCCUPATIONAL_SERIES_CODE_PACKAGE",
    "OPM_PAY_PLAN_CODES",
    "OPM_PAY_PLAN_CODES_2026_08_03",
    "OPM_PAY_PLAN_CODE_PACKAGE",
    "OPM_PLUM_ALL_DATA_URL",
    "OPM_PLUM_DATA_URL",
    "OPM_PLUM_POSITION_STATUS_CODES",
    "OPM_PLUM_POSITION_STATUS_CODES_2026_08_03",
    "OPM_PLUM_POSITION_STATUS_CODE_PACKAGE",
    "OPM_PORTFOLIO_GAPS",
    "OPM_PUBLISHER",
    "OPM_WORKFORCE_DATA_URL",
    "OPM_WORK_SCHEDULE_CODES",
    "OPM_WORK_SCHEDULE_CODES_2026_08_03",
    "OPM_WORK_SCHEDULE_CODE_PACKAGE",
    "AcquiredOPMSource",
    "FetchedOPMResponse",
    "OPMAcquisitionError",
    "OPMAssignmentError",
    "OPMCode",
    "OPMConstantSource",
    "OPMControlPortfolio",
    "OPMControlledListPackageError",
    "OPMControlledListPackageSpec",
    "OPMControlledListView",
    "OPMEHRIDataElement",
    "OPMEHRIDataStandardsExport",
    "OPMEHRIValue",
    "OPMFetcher",
    "OPMFieldAssignment",
    "OPMPLUMAllDataExport",
    "OPMPLUMRow",
    "OPMResourceError",
    "OPMSnapshotPin",
    "OPMSourceDriftError",
    "ParsedOPMResource",
    "ValidatedOPMWorkforceCodes",
    "ValidatedPLUMPositionCodes",
    "acquire_opm_constants",
    "assemble_opm_control_portfolio",
    "build_opm_controlled_list_package",
    "parse_opm_constants",
    "parse_opm_ehri_data_standards_xlsx",
    "parse_opm_plum_all_data_csv",
    "sha256_digest",
    "validate_plum_position_codes",
    "validate_workforce_observation_codes",
]
