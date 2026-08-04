"""Treasury Account Symbol structure and FAST Book edition capture tests."""

from __future__ import annotations

from collections import Counter
from pathlib import Path

import pytest
from openpyxl import load_workbook

from refspec.registry import treasury_tas_fast_book as tas
from refspec.registry.infrastructure.controlled_identifier import ControlledIdentifier

FIXTURES = Path(__file__).parent / "fixtures" / "treasury_tas_fast_book"
TAS_PAGE_FIXTURE = FIXTURES / "treasury-account-symbol-reporting-2026-08-03.html"
FAST_BOOK_PAGE_FIXTURE = FIXTURES / "fast-book-description-of-contents-2026-08-03.html"
FAST_BOOK_WORKBOOK_FIXTURE = FIXTURES / "fast-book-part-ii-iii-2026-07-31.xlsx"
SIZE_AUTHORITY_FIXTURE = FIXTURES / "component-tas-betc-flyer.pdf"


def _acquire(
    tmp_path: Path,
    pin: tas.TreasuryPageSnapshotPin,
    source_path: Path,
) -> tas.AcquiredTreasuryPage:
    return tas.acquire_treasury_page(pin, tmp_path, source_path=source_path)


def test_live_snapshot_pins_match_exact_official_html_bytes() -> None:
    tas_page = TAS_PAGE_FIXTURE.read_bytes()
    fast_book_page = FAST_BOOK_PAGE_FIXTURE.read_bytes()

    assert len(tas_page) == 112_908
    assert tas.sha256_digest(tas_page) == ("sha256:fbd8c6794fdf10d4e1b28ece79af5c15352eb25d292069e0238c3c7513f4675d")
    assert len(fast_book_page) == 110_043
    assert tas.sha256_digest(fast_book_page) == (
        "sha256:91525d80cc4bd6e8ab08075ad630b484b0f691c08516a36151589ddbd57c2a36"
    )
    assert tas.TAS_COMPONENT_FORMAT_2026_08_03.expected_sha256 == tas.sha256_digest(tas_page)
    assert tas.FAST_BOOK_DESCRIPTION_2026_08_03.expected_sha256 == tas.sha256_digest(fast_book_page)


def test_official_fast_book_workbook_parses_all_part_ii_and_iii_rows() -> None:
    payload = FAST_BOOK_WORKBOOK_FIXTURE.read_bytes()

    assert len(payload) == 420_508
    assert tas.sha256_digest(payload) == "sha256:0e40902a2e4bfee7439fbe24d90fd9ff39fad859b4ba432725256866b06cb461"
    assert tas.FAST_BOOK_PART_II_III_2026_07_31.expected_sha256 == tas.sha256_digest(payload)

    parsed = tas.parse_fast_book_workbook(
        FAST_BOOK_WORKBOOK_FIXTURE,
        pin=tas.FAST_BOOK_PART_II_III_2026_07_31,
    )

    assert parsed.edition == "2026-07"
    assert parsed.part_ii_row_count == 3_442
    assert parsed.part_iii_row_count == 140
    assert parsed.change_row_count == 1_159
    assert len(parsed.accounts) == 3_582
    assert len({row.treasury_account_symbol for row in parsed.accounts}) == 3_581
    assert parsed.source_url == "https://tfx.treasury.gov/media/60111/download?inline="
    assert parsed.source_sha256 == tas.FAST_BOOK_PART_II_III_2026_07_31.expected_sha256
    assert parsed.source_byte_length == 420_508
    assert parsed.workbook_modified_at == "2026-07-30T19:11:58"
    assert Counter(row.fund_type for row in parsed.accounts if row.part == "II") == {
        "General Funds": 2_133,
        "Revolving Funds": 377,
        "Special Funds": 376,
        "Trust Funds": 299,
        "Deposit Funds": 246,
        "Consolidated Working Funds": 6,
        "Management Funds": 5,
    }
    assert Counter(row.fund_type for row in parsed.accounts if row.part == "III") == {
        "Program Accounts (not requiring appropriation)": 58,
        "Control Accounts": 41,
        "Holding Accounts": 39,
        "Suspense Accounts": 2,
    }
    assert parsed.accounts[0] == tas.FASTBookPublishedAccount(
        part="II",
        treasury_account_symbol="000 0100",
        agency_identifier="000",
        main_account="0100",
        agency_name="Congress - Senate",
        account_title="Compensation of Members and Related Administrative Expenses, Senate",
        legislation=None,
        fund_type="General Funds",
        independent_agency_identifier=None,
        last_updated="2022-10-26",
    )
    assert parsed.accounts[-1] == tas.FASTBookPublishedAccount(
        part="III",
        treasury_account_symbol="275X7860.308",
        agency_identifier="275",
        main_account="7860.308",
        agency_name="DEPARTMENT OF THE TREASURY",
        account_title=(
            "Japan, United States Friendship Trust Fund, "
            "Japan-United States Friendship Commission, Japan"
        ),
        legislation=None,
        fund_type="Program Accounts (not requiring appropriation)",
        independent_agency_identifier="095",
        last_updated="2024-10-21",
    )

    july_2026 = next(row for row in parsed.accounts if row.treasury_account_symbol == "077 0500")
    assert july_2026.account_title == (
        "Equity Investment Account, United States International Development Finance Corporation"
    )
    assert july_2026.last_updated == "2026-07-01"


def test_fast_book_workbook_reports_publisher_cell_defects_without_rewriting_tas() -> None:
    parsed = tas.parse_fast_book_workbook(
        FAST_BOOK_WORKBOOK_FIXTURE,
        pin=tas.FAST_BOOK_PART_II_III_2026_07_31,
    )

    assert len(parsed.publisher_anomalies) == 6
    assert any("row 911" in item and "X-YEAR" in item for item in parsed.publisher_anomalies)
    assert any("row 1221" in item and "Main cell 'X'" in item for item in parsed.publisher_anomalies)
    assert any("row 1221" in item and "X-YEAR cell 1022" in item for item in parsed.publisher_anomalies)
    assert any("row 687" in item and "Last update is empty" in item for item in parsed.publisher_anomalies)
    assert any("row 1305" in item and "Last update is empty" in item for item in parsed.publisher_anomalies)
    assert any("row 3190" in item and "duplicates Part II row 3185" in item for item in parsed.publisher_anomalies)

    shifted = next(row for row in parsed.accounts if row.treasury_account_symbol == "019X1022")
    assert shifted.agency_identifier == "019"
    assert shifted.main_account == "1022"
    assert shifted.account_title == "International Narcotics Control and Law Enforcement, State"

    identifier = tas.published_fast_book_identifier(
        shifted,
        observed_at=tas.FAST_BOOK_PART_II_III_2026_07_31.retrieved_at,
        source_digest=parsed.source_sha256,
    )
    assert identifier == ControlledIdentifier(
        value="019X1022",
        kind="treasuryAccountSymbol",
        authority_uri=tas.TREASURY_IDENTIFIER_AUTHORITY_URI,
        source_uri=tas.FAST_BOOK_PART_II_III_SOURCE_URL,
        observed_at="2026-08-04T04:36:30Z",
        effective_at="2025-02-27",
        source_digest=tas.FAST_BOOK_PART_II_III_2026_07_31.expected_sha256,
    )


def test_fast_book_workbook_digest_and_headers_fail_closed(tmp_path: Path) -> None:
    changed_bytes = bytearray(FAST_BOOK_WORKBOOK_FIXTURE.read_bytes())
    changed_bytes[-1] ^= 1
    changed_path = tmp_path / "changed.xlsx"
    changed_path.write_bytes(changed_bytes)

    with pytest.raises(tas.TreasurySourceDriftError, match="digest drift"):
        tas.parse_fast_book_workbook(
            changed_path,
            pin=tas.FAST_BOOK_PART_II_III_2026_07_31,
        )

    workbook = load_workbook(FAST_BOOK_WORKBOOK_FIXTURE)
    workbook["Part II"]["A2"] = "Agency Identifier"
    header_changed_path = tmp_path / "header-changed.xlsx"
    workbook.save(header_changed_path)
    saved = load_workbook(header_changed_path, read_only=True, data_only=True)
    assert saved.properties.modified is not None
    header_changed_payload = header_changed_path.read_bytes()
    header_changed_pin = tas.FASTBookWorkbookPin(
        source_url=tas.FAST_BOOK_PART_II_III_SOURCE_URL,
        filename="header-changed.xlsx",
        retrieved_at="2026-08-04T04:36:30Z",
        edition=saved.properties.modified.strftime("%Y-%m"),
        expected_sha256=tas.sha256_digest(header_changed_payload),
        expected_byte_length=len(header_changed_payload),
        expected_modified_at=saved.properties.modified.isoformat(),
        expected_part_ii_rows=3_442,
        expected_part_iii_rows=140,
        expected_change_rows=1_159,
    )

    with pytest.raises(tas.TreasurySourceDriftError, match="Part II headers drifted"):
        tas.parse_fast_book_workbook(header_changed_path, pin=header_changed_pin)


def test_component_size_authority_flyer_matches_its_reference_pin() -> None:
    flyer = SIZE_AUTHORITY_FIXTURE.read_bytes()

    assert len(flyer) == tas.TAS_COMPONENT_SIZE_AUTHORITY_BYTE_LENGTH
    assert tas.sha256_digest(flyer) == tas.TAS_COMPONENT_SIZE_AUTHORITY_SHA256
    # Reference-only provenance: the widths the flyer states are the widths
    # the module's component patterns enforce.
    assert tas._SP_PATTERN.pattern == r"^\d{2}$"
    assert tas._ATA_PATTERN.pattern == r"^\d{3}$"
    assert tas._AID_PATTERN.pattern == r"^\d{3}$"
    assert tas._POA_PATTERN.pattern == r"^\d{4}$"
    assert tas._AVAILABILITY_TYPE_PATTERN.pattern == r"^[A-Z]$"
    assert tas._MAIN_PATTERN.pattern == r"^\d{4}$"
    assert tas._SUB_PATTERN.pattern == r"^\d{3}$"


def test_local_capture_is_content_addressed_and_rechecked_on_cache_hit(
    tmp_path: Path,
) -> None:
    pin = tas.TAS_COMPONENT_FORMAT_2026_08_03

    acquired = _acquire(tmp_path, pin, TAS_PAGE_FIXTURE)
    cached = tas.acquire_treasury_page(pin, tmp_path)

    assert acquired.path == (tmp_path / "sha256" / pin.expected_sha256.removeprefix("sha256:") / pin.source.filename)
    assert acquired.acquisition_mode == "local"
    assert acquired.cache_hit is False
    assert cached.sha256 == pin.expected_sha256
    assert cached.acquisition_mode == "cache"
    assert cached.cache_hit is True


def test_injected_fetcher_is_the_only_live_transport_boundary(tmp_path: Path) -> None:
    payload = FAST_BOOK_PAGE_FIXTURE.read_bytes()
    calls: list[tuple[str, float]] = []

    class Fetcher:
        def fetch(
            self,
            source_url: str,
            *,
            timeout_seconds: float,
        ) -> tas.FetchedTreasuryPage:
            calls.append((source_url, timeout_seconds))
            return tas.FetchedTreasuryPage(
                body=payload,
                status_code=200,
                content_type="text/html; charset=UTF-8",
                resolved_url=source_url,
            )

    acquired = tas.acquire_treasury_page(
        tas.FAST_BOOK_DESCRIPTION_2026_08_03,
        tmp_path,
        fetcher=Fetcher(),
        timeout_seconds=17.0,
    )

    assert calls == [(tas.FAST_BOOK_DESCRIPTION_SOURCE.source_url, 17.0)]
    assert acquired.acquisition_mode == "fetcher"


def test_fetcher_rejects_off_domain_resolved_url(tmp_path: Path) -> None:
    payload = TAS_PAGE_FIXTURE.read_bytes()

    class RedirectedFetcher:
        def fetch(
            self,
            source_url: str,
            *,
            timeout_seconds: float,
        ) -> tas.FetchedTreasuryPage:
            del timeout_seconds
            return tas.FetchedTreasuryPage(
                body=payload,
                status_code=200,
                content_type="text/html",
                resolved_url="https://evil.example/phish",
            )

    with pytest.raises(tas.TreasuryAcquisitionError, match="official HTTPS fiscal.treasury.gov"):
        tas.acquire_treasury_page(
            tas.TAS_COMPONENT_FORMAT_2026_08_03,
            tmp_path,
            fetcher=RedirectedFetcher(),
        )


def test_tas_component_page_parses_documented_field_names_and_edition(
    tmp_path: Path,
) -> None:
    acquired = _acquire(tmp_path, tas.TAS_COMPONENT_FORMAT_2026_08_03, TAS_PAGE_FIXTURE)

    parsed = tas.parse_tas_component_page(acquired)

    assert parsed.edition_date == "2026-03-25"
    assert parsed.source_sha256 == tas.TAS_COMPONENT_FORMAT_2026_08_03.expected_sha256
    assert parsed.component_field_labels == (
        "sub-level prefix (SP)",
        "allocation transfer identifier (ATA)",
        "agency identifier (AID)",
        "beginning period of availability (BPOA)",
        "ending period of availability (EPOA)",
        "availability type (A)",
        "main account (main)",
        "sub-account code (SUB)",
    )
    assert any("does not publish" in gap for gap in parsed.gaps)


def test_fast_book_page_parses_documented_part_fund_groups_and_edition(
    tmp_path: Path,
) -> None:
    acquired = _acquire(tmp_path, tas.FAST_BOOK_DESCRIPTION_2026_08_03, FAST_BOOK_PAGE_FIXTURE)

    parsed = tas.parse_fast_book_description_page(acquired)

    assert parsed.edition_date == "2026-04-29"
    assert parsed.part_fund_groups == {
        "I": ("general", "special", "trust"),
        "II": ("general", "revolving", "special", "deposit", "trust"),
        "III": ("foreignCurrency",),
    }
    assert any("Part II and III" in gap and "parses in full" in gap for gap in parsed.gaps)
    assert any("Part I" in gap and "not included" in gap for gap in parsed.gaps)


def test_digest_or_marker_drift_never_becomes_a_parsed_page(tmp_path: Path) -> None:
    payload = TAS_PAGE_FIXTURE.read_bytes()
    changed = payload.replace(b"agency identifier (AID)", b"agency identifier (AI0)")
    assert len(changed) == len(payload)

    class ChangedFetcher:
        def fetch(
            self,
            source_url: str,
            *,
            timeout_seconds: float,
        ) -> tas.FetchedTreasuryPage:
            del timeout_seconds
            return tas.FetchedTreasuryPage(
                body=changed,
                status_code=200,
                content_type="text/html",
                resolved_url=source_url,
            )

    with pytest.raises(tas.TreasurySourceDriftError, match="digest drift"):
        tas.acquire_treasury_page(
            tas.TAS_COMPONENT_FORMAT_2026_08_03,
            tmp_path,
            fetcher=ChangedFetcher(),
        )

    mini_payload = b"<html><body><p>No component fields here.</p></body></html>"
    mini_pin = tas.TreasuryPageSnapshotPin(
        source=tas.TAS_COMPONENT_FORMAT_SOURCE,
        retrieved_at="2026-08-03T19:17:15Z",
        expected_sha256=tas.sha256_digest(mini_payload),
        expected_byte_length=len(mini_payload),
    )
    mini_path = tmp_path / "mini.html"
    mini_path.write_bytes(mini_payload)
    acquired = tas.acquire_treasury_page(mini_pin, tmp_path / "shape", source_path=mini_path)
    with pytest.raises(tas.TreasurySourceDriftError, match="missing expected component marker"):
        tas.parse_tas_component_page(acquired)


def _record(**changes: str | None) -> dict[str, str | None]:
    values: dict[str, str | None] = {
        "SP": None,
        "ATA": None,
        "AID": "020",
        "BPOA": None,
        "EPOA": None,
        "A": None,
        "MAIN": "0100",
        "SUB": None,
    }
    values.update(changes)
    return values


def test_tas_components_parse_minimal_no_year_account() -> None:
    components = tas.parse_tas_components(_record())

    assert components.sub_level_prefix is None
    assert components.allocation_transfer_agency is None
    assert components.agency_identifier == "020"
    assert components.beginning_period_of_availability is None
    assert components.ending_period_of_availability is None
    assert components.availability_type_code is None
    assert components.main_account == "0100"
    assert components.sub_account == "000"


def test_tas_components_parse_full_multiyear_allocation_account() -> None:
    components = tas.parse_tas_components(
        _record(
            SP="01",
            ATA="012",
            BPOA="2024",
            EPOA="2025",
            A="X",
            SUB="001",
        )
    )

    assert components.sub_level_prefix == "01"
    assert components.allocation_transfer_agency == "012"
    assert components.beginning_period_of_availability == "2024"
    assert components.ending_period_of_availability == "2025"
    assert components.availability_type_code == "X"
    assert components.sub_account == "001"


@pytest.mark.parametrize(
    ("changes", "message"),
    [
        ({"AID": "20"}, "agency_identifier must be exactly 3 digits"),
        ({"AID": "02A"}, "agency_identifier must be exactly 3 digits"),
        ({"MAIN": "100"}, "main_account must be exactly 4 digits"),
        ({"ATA": "12"}, "allocation_transfer_agency must be exactly 3 digits"),
        ({"SUB": "01"}, "sub_account must be exactly 3 digits"),
        ({"SP": "1"}, "sub_level_prefix must be exactly 2 digits"),
        ({"A": "XY"}, "availability_type_code must be exactly one uppercase letter"),
        ({"BPOA": "2024"}, "beginning_period_of_availability and ending_period_of_availability"),
        ({"EPOA": "2024"}, "beginning_period_of_availability and ending_period_of_availability"),
        ({"BPOA": "2025", "EPOA": "2024"}, "must not be after"),
    ],
)
def test_tas_components_reject_malformed_or_inconsistent_fields(
    changes: dict[str, str],
    message: str,
) -> None:
    with pytest.raises(tas.TASComponentError, match=message):
        tas.parse_tas_components(_record(**changes))


def test_tas_components_reject_missing_required_fields() -> None:
    record = _record()
    del record["AID"]
    with pytest.raises(tas.TASComponentError, match="AID"):
        tas.parse_tas_components(record)


def test_tas_components_reject_unknown_fields() -> None:
    with pytest.raises(tas.TASComponentError, match="unknown"):
        tas.parse_tas_components({**_record(), "EXTRA": "1"})


def test_tas_identifier_builds_a_deterministic_capture_local_value() -> None:
    components = tas.parse_tas_components(_record(ATA="012", BPOA="2024", EPOA="2025", A="X"))

    identifier = tas.tas_identifier(
        components,
        observed_at="2026-08-03T19:17:15Z",
        source_digest=tas.TAS_COMPONENT_FORMAT_2026_08_03.expected_sha256,
    )

    assert identifier == ControlledIdentifier(
        value=".012.020.2024.2025.X.0100.000",
        kind="treasuryAccountSymbolComponents",
        authority_uri=tas.TREASURY_IDENTIFIER_AUTHORITY_URI,
        source_uri=tas.TAS_COMPONENT_FORMAT_SOURCE.source_url,
        observed_at="2026-08-03T19:17:15Z",
        effective_at=None,
        source_digest=tas.TAS_COMPONENT_FORMAT_2026_08_03.expected_sha256,
    )
    # Parsing the canonical value back must reproduce every component exactly.
    assert tas.parse_tas_canonical_value(identifier.value) == components


def test_fast_book_account_record_validates_against_its_parsed_part_fund_groups(
    tmp_path: Path,
) -> None:
    description = tas.parse_fast_book_description_page(
        _acquire(tmp_path, tas.FAST_BOOK_DESCRIPTION_2026_08_03, FAST_BOOK_PAGE_FIXTURE)
    )

    record = tas.validate_fast_book_account_record(
        {
            "AID": "020",
            "MAIN": "0100",
            "ACCOUNT_TITLE": "Example Trust Fund",
            "PART": "I",
            "FUND_GROUP": "trust",
            "STATUTORY_CITATION": "31 U.S.C. 1321",
        },
        description=description,
    )

    assert record.agency_identifier == "020"
    assert record.main_account == "0100"
    assert record.account_title == "Example Trust Fund"
    assert record.fast_book_part == "I"
    assert record.fund_group == "trust"
    assert record.statutory_citation == "31 U.S.C. 1321"
    assert record.edition_date == "2026-04-29"


def test_fast_book_account_record_rejects_fund_group_outside_its_part(
    tmp_path: Path,
) -> None:
    description = tas.parse_fast_book_description_page(
        _acquire(tmp_path, tas.FAST_BOOK_DESCRIPTION_2026_08_03, FAST_BOOK_PAGE_FIXTURE)
    )

    with pytest.raises(tas.FASTBookRecordError, match="not a documented fund group for Part I"):
        tas.validate_fast_book_account_record(
            {
                "AID": "020",
                "MAIN": "0100",
                "ACCOUNT_TITLE": "Example Deposit Fund",
                "PART": "I",
                "FUND_GROUP": "deposit",
                "STATUTORY_CITATION": None,
            },
            description=description,
        )


def test_fast_book_account_record_requires_citation_for_special_and_trust_funds(
    tmp_path: Path,
) -> None:
    description = tas.parse_fast_book_description_page(
        _acquire(tmp_path, tas.FAST_BOOK_DESCRIPTION_2026_08_03, FAST_BOOK_PAGE_FIXTURE)
    )

    with pytest.raises(tas.FASTBookRecordError, match="requires a statutory citation"):
        tas.validate_fast_book_account_record(
            {
                "AID": "020",
                "MAIN": "0100",
                "ACCOUNT_TITLE": "Example Trust Fund",
                "PART": "I",
                "FUND_GROUP": "trust",
                "STATUTORY_CITATION": None,
            },
            description=description,
        )

    with pytest.raises(tas.FASTBookRecordError, match="general fund account must not carry a statutory citation"):
        tas.validate_fast_book_account_record(
            {
                "AID": "020",
                "MAIN": "0100",
                "ACCOUNT_TITLE": "Example General Fund Account",
                "PART": "I",
                "FUND_GROUP": "general",
                "STATUTORY_CITATION": "31 U.S.C. 1321",
            },
            description=description,
        )


def test_fast_book_identifier_builds_a_deterministic_capture_local_value(
    tmp_path: Path,
) -> None:
    description = tas.parse_fast_book_description_page(
        _acquire(tmp_path, tas.FAST_BOOK_DESCRIPTION_2026_08_03, FAST_BOOK_PAGE_FIXTURE)
    )
    record = tas.validate_fast_book_account_record(
        {
            "AID": "020",
            "MAIN": "0100",
            "ACCOUNT_TITLE": "Example Trust Fund",
            "PART": "I",
            "FUND_GROUP": "trust",
            "STATUTORY_CITATION": "31 U.S.C. 1321",
        },
        description=description,
    )

    identifier = tas.fast_book_identifier(
        record,
        observed_at="2026-08-03T19:17:43Z",
        source_digest=tas.FAST_BOOK_DESCRIPTION_2026_08_03.expected_sha256,
    )

    assert identifier.value == "020-0100"
    assert identifier.kind == "fastBookAccountIdentifier"
    assert identifier.authority_uri == tas.TREASURY_IDENTIFIER_AUTHORITY_URI
    assert identifier.source_uri == tas.FAST_BOOK_DESCRIPTION_SOURCE.source_url


def test_edition_combines_both_pages_and_never_claims_a_universal_page_hash(
    tmp_path: Path,
) -> None:
    tas_parsed = tas.parse_tas_component_page(_acquire(tmp_path, tas.TAS_COMPONENT_FORMAT_2026_08_03, TAS_PAGE_FIXTURE))
    fast_book_parsed = tas.parse_fast_book_description_page(
        _acquire(tmp_path, tas.FAST_BOOK_DESCRIPTION_2026_08_03, FAST_BOOK_PAGE_FIXTURE)
    )

    edition = tas.assemble_treasury_tas_fast_book_edition(tas_parsed, fast_book_parsed)

    assert edition.tas_component_format_edition == "2026-03-25"
    assert edition.fast_book_description_edition == "2026-04-29"
    assert any("Akamai" in gap or "analytics" in gap for gap in edition.gaps)
    assert any("Part II and III" in gap and "parses in full" in gap for gap in edition.gaps)
    assert any("Part I" in gap and "not included" in gap for gap in edition.gaps)
